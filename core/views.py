from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from .models import ( Produto, Bar, Restaurante, RequisicaoProduto, TransferenciaBar, ContagemBar, EstoqueBar, models,
 AcessoUsuarioBar, EventoProduto, Evento, PermissaoPagina, RecebimentoEstoque, EventoAlimento, Alimento, PerdaProduto) 
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from itertools import zip_longest
from datetime import time
from django.core.paginator import Paginator
from django.db import transaction
from django.core.exceptions import FieldError
import uuid
import pandas as pd
import openpyxl
import io
from io import BytesIO
import xlsxwriter
from django.utils.timezone import is_aware, localtime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from django.utils import timezone
from django.db.models import Sum, Max, Q, Count
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.decorators import user_passes_test
from django.db.models import DateField
from django.core.files.storage import default_storage
from django.db.models.functions import TruncDate
from collections import defaultdict, OrderedDict
from django.utils.timezone import now
from django.utils.dateparse import parse_date
from babel.dates import parse_date
from datetime import datetime, date
from django.utils.text import slugify

def login_view(request):
    context = {}
    # suporta ?next=/alguma-rota/
    next_url = request.GET.get('next') or request.POST.get('next')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Só exigiremos restaurante se NÃO houver next (ou seja, fluxo padrão)
        restaurante_id = (request.POST.get('restaurante') or '').strip()

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            # Se veio next, honramos e vamos direto (ex.: página de eventos)
            if next_url:
                return redirect(next_url)

            # Fluxo antigo: exige restaurante válido e acesso, seta na sessão e vai para selecionar-bar
            if not restaurante_id:
                messages.error(request, 'Selecione um restaurante.')
                return redirect('login')

            acesso_valido = AcessoUsuarioBar.objects.filter(user=user, restaurante_id=restaurante_id).exists()
            if not acesso_valido:
                messages.error(request, 'Você não tem acesso a este restaurante.')
                return redirect('login')

            request.session['restaurante_id'] = restaurante_id
            return redirect('selecionar-bar')

        else:
            context['erro'] = 'Usuário ou senha inválidos'

    context['restaurantes'] = Restaurante.objects.all()  # dropdown
    # mantém o next na página (para o botão direto)
    if next_url:
        context['next'] = next_url
    return render(request, 'core/login.html', context)


def logout_view(request):
    logout(request)

    # Limpa manualmente dados da sessão, se necessário
    request.session.flush()

    return redirect('login')


@login_required
def selecionar_bar_view(request):
    # Restaurantes em que o usuário possui acesso
    acessos_user = (AcessoUsuarioBar.objects
                    .filter(user=request.user)
                    .select_related('restaurante'))

    restaurantes_permitidos = (Restaurante.objects
                               .filter(id__in=acessos_user.values_list('restaurante_id', flat=True).distinct())
                               .order_by('nome'))

    # Trocar restaurante explicitamente
    if request.method == 'POST' and request.POST.get('acao') == 'trocar_restaurante':
        request.session.pop('restaurante_id', None)
        request.session.pop('bar_id', None)
        request.session.pop('bar_nome', None)
        return redirect('selecionar-bar')

    restaurante_id = request.session.get('restaurante_id')

    # --------- ETAPA 1: selecionar restaurante ----------
    if not restaurante_id:
        if request.method == 'POST':
            rid = (request.POST.get('restaurante') or '').strip()
            if rid and restaurantes_permitidos.filter(id=rid).exists():
                request.session['restaurante_id'] = int(rid)
                request.session.pop('bar_id', None)
                request.session.pop('bar_nome', None)
                return redirect('selecionar-bar')
            messages.error(request, "Selecione um restaurante válido.")

        return render(request, 'core/selecionar_bar.html', {
            'etapa': 'restaurante',
            'restaurantes': restaurantes_permitidos,
        })

    # --------- ETAPA 2: selecionar bar ----------
    restaurante = get_object_or_404(Restaurante, id=restaurante_id)

    bares_ids = (acessos_user
                 .filter(restaurante_id=restaurante_id)
                 .values_list('bares__id', flat=True))
    bares = Bar.objects.filter(id__in=bares_ids).order_by('nome')

    if request.method == 'POST':
        bar_id = (request.POST.get('bar') or '').strip()
        if bar_id and bares.filter(id=bar_id).exists():
            bar = Bar.objects.get(id=bar_id)
            request.session['bar_id'] = bar.id
            request.session['bar_nome'] = bar.nome
            next_url = request.GET.get('next')
            return redirect(next_url or 'dashboard')
        messages.error(request, "Bar inválido ou sem permissão.")

    return render(request, 'core/selecionar_bar.html', {
        'etapa': 'bar',
        'restaurante': restaurante,
        'bares': bares,
        'restaurantes': restaurantes_permitidos,  # opcional: para combo de troca rápida
    })


@login_required
def requisicao_produtos_view(request):
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='requisicoes').exists():
        messages.error(request, 'Você não tem permissão para acessar essa página.')
        return redirect('dashboard')

    if request.method == 'POST':
        produtos = request.POST.getlist('produto[]')
        quantidades = request.POST.getlist('quantidade[]')

        restaurante_id = request.session.get('restaurante_id')
        bar_id = request.session.get('bar_id')

        restaurante = Restaurante.objects.get(id=restaurante_id)
        bar = Bar.objects.get(id=bar_id)

        try:
            estoque_central = Bar.objects.get(restaurante=restaurante, is_estoque_central=True)
        except Bar.DoesNotExist:
            messages.error(request, 'Estoque central não encontrado para o restaurante.')
            return redirect('requisicao')

        produtos_requisitados = []
        erros = []

        for prod_id, qtd in zip(produtos, quantidades):
            produto = Produto.objects.get(id=prod_id)

            try:
                qtd_solicitada = Decimal(qtd.replace(",", ".")) if qtd else Decimal("0")
            except:
                qtd_solicitada = Decimal("0")

            estoque = EstoqueBar.objects.filter(bar=estoque_central, produto=produto).first()

            if not estoque or estoque.quantidade_garrafas < qtd_solicitada:
                erros.append(f"❌ {produto.nome} - Estoque insuficiente no estoque central.")
                continue

            # Cria a requisição
            RequisicaoProduto.objects.create(
                restaurante=restaurante,
                bar=bar,  # ← bar solicitante
                produto=produto,
                quantidade_solicitada=qtd_solicitada,
                usuario=request.user
            )
            produtos_requisitados.append(produto.nome)

        if produtos_requisitados:
            messages.success(
                request,
                f"Requisição enviada para: {', '.join(produtos_requisitados)}"
            )

        if erros:
            for erro in erros:
                messages.error(request, erro)

        return redirect('requisicao')

    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    return render(request, 'core/requisicao.html', {'produtos': produtos})




@login_required
def entrada_mercadorias_view(request):
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='entrada_mercadoria').exists():
        messages.error(request, 'Você não tem permissão para acessar essa página.')
        return redirect('dashboard')

    restaurante_id = request.session.get('restaurante_id')
    restaurante = Restaurante.objects.get(id=restaurante_id)

    try:
        estoque_central = Bar.objects.get(restaurante=restaurante, is_estoque_central=True)
    except Bar.DoesNotExist:
        messages.error(request, "Estoque central não definido para este restaurante.")
        return redirect('dashboard')

    if request.method == 'POST':
        produtos = request.POST.getlist('produto[]')
        quantidades = request.POST.getlist('quantidade[]')

        for prod_id, qtd in zip(produtos, quantidades):
            produto = Produto.objects.get(id=prod_id)
            quantidade = Decimal((qtd or '').replace(',', '.')) if qtd else Decimal(0)

            # Monta kwargs para não quebrar caso o modelo ainda não tenha esses campos
            kwargs = dict(
                restaurante=restaurante,
                bar=estoque_central,
                produto=produto,
                quantidade=quantidade
            )
            # Se o modelo tiver usuario/data_recebimento/observacao, preenche
            if hasattr(RecebimentoEstoque, 'usuario'):
                kwargs['usuario'] = request.user
            if hasattr(RecebimentoEstoque, 'data_recebimento'):
                kwargs['data_recebimento'] = timezone.now()
            # Observação opcional vinda do form (ex.: um <input name="observacao"> geral ou por linha)
            obs = (request.POST.get('observacao') or '').strip()
            if obs and hasattr(RecebimentoEstoque, 'observacao'):
                kwargs['observacao'] = obs

            # 1) Registra a entrada
            RecebimentoEstoque.objects.create(**kwargs)

            # 2) Atualiza estoque do bar central
            EstoqueBar.adicionar(estoque_central, produto, quantidade)

        messages.success(request, "Entrada de mercadorias realizada com sucesso!")
        return redirect('entrada-mercadorias')

    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    return render(request, 'core/entrada_mercadorias.html', {'produtos': produtos})



@login_required
def historico_entradas_view(request):
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='historico_entrada').exists():
        messages.error(request, "Você não tem permissão para acessar o Histórico de Entradas.")
        return redirect('dashboard')

    restaurante_id = request.session.get('restaurante_id')
    if not restaurante_id:
        messages.error(request, "Restaurante não selecionado.")
        return redirect('dashboard')

    # Base
    entradas = RecebimentoEstoque.objects.filter(restaurante_id=restaurante_id).select_related(
        'produto', 'bar', 'usuario'
    )

    # Campo de data preferencial
    data_field = 'data_recebimento' if hasattr(RecebimentoEstoque, 'data_recebimento') else (
                 'created_at' if hasattr(RecebimentoEstoque, 'created_at') else None)

    # Ordenação alinhada ao índice
    if data_field:
        entradas = entradas.order_by(f'-{data_field}', '-id')
    else:
        entradas = entradas.order_by('-id')  # fallback

    # Filtros
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    before = request.GET.get('before')  # cursor para datas anteriores (YYYY-MM-DD)

    # Filtro Mês/Ano: preferir range (aproveita índice)
    if mes and ano and data_field:
        ano = int(ano); mes = int(mes)
        inicio = date(ano, mes, 1)
        if mes == 12:
            fim = date(ano + 1, 1, 1)
        else:
            fim = date(ano, mes + 1, 1)
        entradas = entradas.filter(**{f'{data_field}__gte': inicio, f'{data_field}__lt': fim})

    agrupado = {}
    next_before = None

    # Quando NÃO está filtrando por mês/ano, aplicamos o modelo "10 dias + cursor"
    if not (mes and ano):
        if before and data_field:
            # pega blocos de 10 dias anteriores ao dia passado
            entradas = entradas.filter(**{f'{data_field}__date__lt': before})

        # subquery leve só com as 10 últimas datas distintas
        if data_field:
            ultimos_dias = (entradas
                            .annotate(d=TruncDate(data_field))
                            .values_list('d', flat=True)
                            .distinct()
                            .order_by('-d')[:10])
            dias = list(ultimos_dias)
            if dias:
                entradas = (entradas
                            .filter(**{f'{data_field}__date__in': dias})
                            .only('id', 'quantidade',
                                  'produto__nome', 'bar__nome',
                                  'usuario__username', data_field))
                # agrupa em memória
                tmp = defaultdict(list)
                for e in entradas:
                    dia = getattr(e, data_field).date() if data_field else 'Sem data'
                    tmp[dia].append(e)
                # ordena as chaves (decrescente)
                agrupado = dict(sorted(tmp.items(), key=lambda kv: kv[0], reverse=True))
                next_before = min(dias).isoformat()
            else:
                agrupado = {}
        else:
            # sem campo de data (raro) — só lista últimos N
            tmp = defaultdict(list)
            for e in entradas[:200]:
                tmp['Sem data'].append(e)
            agrupado = dict(tmp)
    else:
        # Mês/Ano: lista tudo do período (a consulta já veio estreita pelo índice)
        if data_field:
            entradas = entradas.annotate(d=TruncDate(data_field))
            tmp = defaultdict(list)
            for e in entradas:
                tmp[e.d].append(e)
            agrupado = dict(sorted(tmp.items(), key=lambda kv: kv[0], reverse=True))
        else:
            tmp = defaultdict(list)
            for e in entradas:
                tmp['Sem data'].append(e)
            agrupado = dict(tmp)

    return render(request, 'core/historico_entradas.html', {
        'agrupado': agrupado,
        'now': timezone.now(),
        'meses': list(range(1, 13)),
        'data_field': data_field,
        'next_before': next_before,  # para o botão "Ver dias anteriores"
    })




@login_required
def contagem_view(request):
    # 🔒 Permissão
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='contagem').exists():
        messages.error(request, "Você não tem permissão para acessar a página de contagem.")
        return redirect('dashboard')

    bar_id = request.session.get('bar_id')
    restaurante_id = request.session.get('restaurante_id')
    bar = Bar.objects.get(id=bar_id)
    restaurante = Restaurante.objects.get(id=restaurante_id)

    # Produtos que "existem" nesse bar (via EstoqueBar). Fallback: todos ativos
    produtos_ids_no_bar = list(
        EstoqueBar.objects.filter(bar=bar).values_list('produto_id', flat=True)
    )
    if produtos_ids_no_bar:
        produtos = Produto.objects.filter(ativo=True, id__in=produtos_ids_no_bar).order_by('nome')
    else:
        produtos = Produto.objects.filter(ativo=True).order_by('nome')

    # Mapa de estoque atual para pré-preenchimento
    estoques = EstoqueBar.objects.filter(bar=bar, produto__in=produtos)
    estoque_map = {e.produto_id: e for e in estoques}

    if request.method == 'POST':
        # Percorre SEMPRE a lista fixa que renderizamos
        for p in produtos:
            g_raw = (request.POST.get(f'garrafas_{p.id}', '') or '').strip()
            d_raw = (request.POST.get(f'doses_{p.id}', '') or '').strip()

            # Se usuário não mexeu em nada nessa linha, ignore
            if g_raw == '' and d_raw == '':
                continue

            # Sanitização
            try:
                g_val = int(g_raw) if g_raw != '' else 0
            except ValueError:
                g_val = 0

            try:
                d_val = Decimal(d_raw.replace(',', '.')) if d_raw != '' else Decimal('0')
            except (InvalidOperation, ValueError):
                d_val = Decimal('0')

            # 1) Registrar contagem
            ContagemBar.objects.create(
                bar=bar,
                produto=p,
                quantidade_garrafas_cheias=g_val,
                quantidade_doses_restantes=d_val,
                usuario=request.user,
                data_contagem=timezone.now()
            )

            # 2) Atualizar estoque atual
            estoque, _ = EstoqueBar.objects.get_or_create(bar=bar, produto=p)
            estoque.quantidade_garrafas = g_val
            estoque.quantidade_doses = d_val
            estoque.save()

        messages.success(request, "Contagem registrada e estoque atualizado com sucesso!")
        return redirect('contagem')

    # Monta linhas para o template (com último valor como placeholder)
    linhas = []
    for p in produtos:
        e = estoque_map.get(p.id)
        linhas.append({
            'produto': p,
            'g_prev': (e.quantidade_garrafas if e else 0),
            'd_prev': (e.quantidade_doses if e else Decimal('0')),
        })

    return render(request, 'core/contagem.html', {
        'bar': bar,
        'linhas': linhas,
    })




@login_required
def historico_contagens_view(request):
    # 🔒 Verificação de permissão para "historico_contagens"
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='historico_cont').exists():
        messages.error(request, "Você não tem permissão para acessar a página de Historico de contagens.")
        return redirect('dashboard')
    
    bar_id = request.session.get('bar_id')
    bar = Bar.objects.get(id=bar_id)

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    contagens = ContagemBar.objects.filter(bar=bar)

    if mes and ano:
        contagens = contagens.filter(data_contagem__month=mes, data_contagem__year=ano)
    else:
        contagens = contagens.annotate(data=TruncDate('data_contagem')).order_by('-data_contagem')

    agrupado = {}
    for c in contagens:
        data = c.data_contagem.date()
        if mes and ano:
            agrupado.setdefault(data, []).append(c)
        else:
            if data not in agrupado:
                if len(agrupado) >= 10:
                    break
                agrupado[data] = []
            agrupado[data].append(c)

    return render(request, 'core/historico_contagens.html', {
        'agrupado': agrupado,
        'now': now(),
        'meses': list(range(1, 13))  # de 1 a 12
    })





@login_required
def aprovar_requisicoes_view(request):
    # 🔒 Permissão
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='aprovacao').exists():
        messages.error(request, "Você não tem permissão para acessar a página de Aprovar/Rejeitar.")
        return redirect('dashboard')

    restaurante_id = request.session.get('restaurante_id')
    requisicoes = RequisicaoProduto.objects.filter(restaurante_id=restaurante_id, status='PENDENTE')
    restaurante = get_object_or_404(Restaurante, id=restaurante_id)
    bar_central = get_object_or_404(Bar, restaurante=restaurante, is_estoque_central=True)

    if request.method == 'POST':
        erros = []
        # percorre apenas as linhas que receberam decisão
        for key in request.POST:
            if not key.startswith('aprovacao_'):
                continue

            req_id = key.split('_')[1]
            decisao = request.POST.get(key)
            requisicao = get_object_or_404(
                RequisicaoProduto, id=req_id, restaurante_id=restaurante_id, status='PENDENTE'
            )

            if decisao == 'aprovar':
                qtd = Decimal(requisicao.quantidade_solicitada)
                sucesso = EstoqueBar.retirar(bar_central, requisicao.produto, qtd)

                if sucesso:
                    EstoqueBar.adicionar(requisicao.bar, requisicao.produto, qtd)

                    TransferenciaBar.objects.create(
                        restaurante=requisicao.restaurante,
                        origem=bar_central,
                        destino=requisicao.bar,
                        produto=requisicao.produto,
                        quantidade=qtd,
                        usuario=request.user
                    )

                    requisicao.status = 'APROVADA'
                    messages.success(request, "Requisição aprovada com sucesso.")
                else:
                    requisicao.status = 'FALHA_ESTOQUE'
                    requisicao.motivo_negativa = "Produto insuficiente no estoque central."
                    messages.warning(
                        request,
                        f"Produto '{requisicao.produto.nome}' insuficiente no estoque central. Requisição {req_id} não aprovada."
                    )

                requisicao.usuario_aprovador = request.user
                requisicao.data_decisao = timezone.now()
                requisicao.save()

            elif decisao == 'negar':
                motivo = (request.POST.get(f'motivo_{req_id}', '') or '').strip()
                if not motivo:
                    erros.append(f"Informe o motivo da negativa para a requisição {req_id}.")
                    # não salva/avança essa requisição — continua PENDENTE
                    continue

                requisicao.status = 'NEGADA'
                requisicao.motivo_negativa = motivo
                requisicao.usuario_aprovador = request.user
                requisicao.data_decisao = timezone.now()
                requisicao.save()
                messages.info(request, f"Requisição {req_id} negada.")

        if erros:
            for e in erros:
                messages.error(request, e)
        return redirect('aprovar-requisicoes')

    return render(request, 'core/aprovar_requisicoes.html', {'requisicoes': requisicoes})






def atualizar_estoque(bar, produto, quantidade_delta):
    estoque, _ = EstoqueBar.objects.get_or_create(bar=bar, produto=produto)
    estoque.quantidade += quantidade_delta
    estoque.save()



@login_required
def historico_requisicoes_view(request):
    # 🔒 Permissão
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='historico_requi').exists():
        messages.error(request, "Você não tem permissão para acessar a página de Historico de requisições.")
        return redirect('dashboard')

    bar_id = request.session.get('bar_id')
    restaurante_id = request.session.get('restaurante_id')

    filtro_ativo = False
    agrupado = defaultdict(list)

    if not bar_id or not restaurante_id:
        return render(request, 'core/historico_requisicoes.html', {
            'agrupado': agrupado,
            'now': datetime.now(),
            'filtro_ativo': filtro_ativo
        })

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    # ⚡ mais performático para render (tudo que exibimos é FK direta)
    qs = (RequisicaoProduto.objects
          .filter(restaurante_id=restaurante_id, bar_id=bar_id)
          .select_related('produto', 'bar', 'usuario', 'usuario_aprovador'))

    if mes and ano:
        try:
            mes_i = int(mes)
            ano_i = int(ano)
            qs = qs.filter(data_solicitacao__month=mes_i, data_solicitacao__year=ano_i)
            filtro_ativo = True
        except ValueError:
            pass
        qs = qs.order_by('-data_solicitacao')
    else:
        # Últimas 20 (sem filtro)
        qs = qs.order_by('-data_solicitacao')[:20]

    # Agrupa por data (dia)
    qs = qs.annotate(data_truncada=TruncDate('data_solicitacao'))
    for r in qs:
        agrupado[r.data_truncada].append(r)

    return render(request, 'core/historico_requisicoes.html', {
        'agrupado': dict(agrupado),
        'now': datetime.now(),
        'filtro_ativo': filtro_ativo
    })





from decimal import Decimal, InvalidOperation

@login_required
def transferencia_entre_bares_view(request):
    # 🔒 Verificação de permissão
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='transferencias').exists():
        messages.error(request, "Você não tem permissão para acessar a página de transferências.")
        return redirect('dashboard')

    restaurante_id = request.session.get('restaurante_id')
    bar_origem_id = request.session.get('bar_id')

    restaurante = Restaurante.objects.get(id=restaurante_id)
    bar_origem = Bar.objects.get(id=bar_origem_id)
    bares_destino = Bar.objects.filter(restaurante=restaurante).exclude(id=bar_origem_id)

    if request.method == 'POST':
        produto_id = request.POST.get('produto')
        bar_destino_id = request.POST.get('bar_destino')
        quantidade_str = request.POST.get('quantidade')

        # 🔎 Validação dos campos
        if not produto_id or not bar_destino_id or not quantidade_str:
            messages.error(request, "Todos os campos são obrigatórios.")
            return redirect('transferencia-bares')

        # 🔢 Tentar converter a quantidade
        try:
            quantidade = Decimal(quantidade_str.replace(',', '.'))
        except (InvalidOperation, AttributeError):
            messages.error(request, "Quantidade inválida.")
            return redirect('transferencia-bares')

        # 🔎 Verifica se produto e bar de destino existem
        try:
            produto = Produto.objects.get(id=produto_id)
            bar_destino = Bar.objects.get(id=bar_destino_id)
        except Produto.DoesNotExist:
            messages.error(request, "Produto inválido.")
            return redirect('transferencia-bares')
        except Bar.DoesNotExist:
            messages.error(request, "Bar de destino inválido.")
            return redirect('transferencia-bares')

        # 🧮 Verifica se há estoque suficiente no bar de origem
        sucesso = EstoqueBar.retirar(bar_origem, produto, quantidade)

        if not sucesso:
            messages.error(request, "Estoque insuficiente para essa transferência.")
            return redirect('transferencia-bares')

        # ➕ Adiciona ao bar de destino
        EstoqueBar.adicionar(bar_destino, produto, quantidade)

        # 📝 Registra a transferência
        TransferenciaBar.objects.create(
            restaurante=restaurante,
            origem=bar_origem,
            destino=bar_destino,
            produto=produto,
            quantidade=quantidade,
            usuario=request.user
        )

        messages.success(request, "Transferência realizada com sucesso!")
        return redirect('transferencia-bares')

    # 🧾 Carrega os produtos com estoque no bar de origem
    produtos_disponiveis = EstoqueBar.objects.filter(bar=bar_origem, quantidade_garrafas__gt=0)

    return render(request, 'core/transferencia_bares.html', {
        'bares_destino': bares_destino,
        'produtos': [e.produto for e in produtos_disponiveis]
    })






@login_required
def historico_transferencias_view(request):
    # 🔒 Verificação de permissão para "historico_transferencias"
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='historico_transf').exists():
        messages.error(request, "Você não tem permissão para acessar a página de Historico de transferencias.")
        return redirect('dashboard')
    
    bar_id = request.session.get('bar_id')
    restaurante_id = request.session.get('restaurante_id')

    filtro_ativo = False
    agrupado = defaultdict(list)  # ✅ evita KeyError automaticamente

    if not bar_id or not restaurante_id:
        return render(request, 'core/historico_transferencias.html', {
            'agrupado': agrupado,
            'now': datetime.now(),
            'filtro_ativo': filtro_ativo
        })

    bar = Bar.objects.get(id=bar_id)
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    # 🔍 Base query
    transferencias = TransferenciaBar.objects.filter(
        restaurante_id=restaurante_id
    ).filter(
        models.Q(origem=bar) | models.Q(destino=bar)
    )

    if mes and ano:
        try:
            mes = int(mes)
            ano = int(ano)
            transferencias = transferencias.filter(
                data_transferencia__month=mes,
                data_transferencia__year=ano
            )
            filtro_ativo = True
        except ValueError:
            pass
    else:
        transferencias = transferencias.order_by('-data_transferencia')[:10]

    # ✅ Anota data truncada para agrupar
    transferencias = transferencias.annotate(
        data_truncada=TruncDate('data_transferencia')
    )

    # ✅ Agora agrupamos sem risco de erro
    for t in transferencias:
        agrupado[t.data_truncada].append(t)

    return render(request, 'core/historico_transferencias.html', {
        'agrupado': dict(agrupado),
        'now': datetime.now(),
        'filtro_ativo': filtro_ativo
    })


# views.py (trechos relevantes)

from decimal import Decimal, InvalidOperation
from collections import defaultdict
from itertools import zip_longest
from datetime import datetime, time

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from .models import (
    Evento, Produto, Alimento, EventoProduto, EventoAlimento,
    Restaurante, PermissaoPagina
)


@login_required
def pagina_eventos(request):
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='eventos').exists():
        messages.error(request, "Você não tem permissão para acessar a página de eventos.")
        return redirect('dashboard')

    hoje = timezone.localdate()

    # Filtro GET ?restaurante=<id>
    restaurante_param = (request.GET.get('restaurante') or '').strip()
    try:
        restaurante_filtro_id = int(restaurante_param) if restaurante_param else None
    except ValueError:
        restaurante_filtro_id = None

    eventos_qs = (
        Evento.objects
        .exclude(status='FINALIZADO')
        .select_related('restaurante', 'responsavel')
        .prefetch_related('produtos__produto', 'alimentos__alimento')
        .order_by('data_evento', 'data_criacao')
    )
    if restaurante_filtro_id:
        eventos_qs = eventos_qs.filter(restaurante_id=restaurante_filtro_id)

    paginator = Paginator(eventos_qs, 10)
    page_number = request.GET.get("page")
    eventos_abertos = paginator.get_page(page_number)

    # Consolidado: finalizados hoje após 06:00
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(hoje, time(6, 0)), tz)
    end_dt   = timezone.make_aware(datetime.combine(hoje, time(23, 59, 59, 999999)), tz)

    finalizados_qs = (
        Evento.objects.filter(status='FINALIZADO', finalizado_em__range=(start_dt, end_dt))
        .select_related('restaurante')
        .prefetch_related('produtos__produto', 'alimentos__alimento')
        .order_by('-finalizado_em')
    )

    consolidado_bebidas = defaultdict(lambda: {'garrafas': 0, 'doses': 0, 'ml': 0})
    consolidado_alimentos = defaultdict(lambda: {'quantidade': Decimal('0.00'), 'unidade': ''})

    for evento in finalizados_qs:
        for item in evento.produtos.all():
            g = int(item.garrafas or 0)
            d = int(item.doses or 0)
            consolidado_bebidas[item.produto.nome]['garrafas'] += g
            consolidado_bebidas[item.produto.nome]['doses'] += d
            consolidado_bebidas[item.produto.nome]['ml']    += d * 50
        for item in evento.alimentos.all():
            nome = item.alimento.nome
            consolidado_alimentos[nome]['quantidade'] += (item.quantidade or Decimal('0'))
            consolidado_alimentos[nome]['unidade'] = item.alimento.unidade

    produtos = Produto.objects.all().order_by('nome')
    alimentos = Alimento.objects.filter(ativo=True).order_by('nome')
    restaurantes = Restaurante.objects.all().order_by('nome')

    return render(request, 'eventos/pagina_eventos.html', {
        'produtos': produtos,
        'alimentos': alimentos,
        'restaurantes': restaurantes,
        'eventos_abertos': eventos_abertos,
        'paginator': paginator,
        'consolidado_bebidas': dict(consolidado_bebidas),
        'consolidado_alimentos': dict(consolidado_alimentos),
        'hoje': hoje,
        'janela_consolidado_label': "Finalizados hoje após 06:00",
        'selected_restaurante': restaurante_filtro_id,
    })


def _to_int_or_zero(val):
    try:
        return max(int(val), 0)
    except (TypeError, ValueError):
        return 0

def _to_decimal_or_zero(val):
    if val is None:
        return Decimal("0")
    s = str(val).strip().replace(",", ".")
    try:
        q = Decimal(s)
    except (InvalidOperation, ValueError):
        q = Decimal("0")
    if q < 0:
        q = Decimal("0")
    return q


@login_required
@transaction.atomic
def criar_evento(request):
    if request.method != 'POST':
        return redirect('pagina_eventos')

    nome = (request.POST.get('nome_evento') or '').strip()
    pessoas_raw = (request.POST.get('numero_pessoas') or '').strip()
    horas_raw = (request.POST.get('horas') or '').strip()
    data_evento_raw = (request.POST.get('data_evento') or '').strip()
    restaurante_id = (request.POST.get('restaurante_id') or '').strip()

    # numero_pessoas
    try:
        numero_pessoas = int(pessoas_raw) if pessoas_raw != '' else None
        if numero_pessoas is not None and numero_pessoas < 0:
            numero_pessoas = None
    except ValueError:
        numero_pessoas = None

    # horas
    try:
        horas = Decimal(horas_raw.replace(',', '.')) if horas_raw != '' else None
        if horas is not None and horas < 0:
            horas = None
    except Exception:
        horas = None

    # data_evento
    try:
        data_evento = datetime.strptime(data_evento_raw, "%Y-%m-%d").date() if data_evento_raw else timezone.localdate()
    except Exception:
        data_evento = timezone.localdate()

    # restaurante (opcional)
    restaurante = None
    if restaurante_id:
        try:
            restaurante = Restaurante.objects.get(id=restaurante_id)
        except Restaurante.DoesNotExist:
            restaurante = None

    evento = Evento.objects.create(
        nome=nome or f"Evento {timezone.localtime(timezone.now()):%d/%m %H:%M}",
        responsavel=request.user,
        numero_pessoas=numero_pessoas,
        horas=horas,
        status='ABERTO',
        data_evento=data_evento,
        restaurante=restaurante,
    )

    # ---------- Bebidas ----------
    produtos_ids = request.POST.getlist('produto_id[]')
    garrafas_list = request.POST.getlist('garrafas[]')
    doses_list = request.POST.getlist('doses[]')

    soma_produtos = defaultdict(lambda: {'garrafas': 0, 'doses': Decimal('0')})
    for pid, g_raw, d_raw in zip_longest(produtos_ids, garrafas_list, doses_list, fillvalue='0'):
        pid = (pid or '').strip()
        if not pid:
            continue
        g = max(int(g_raw or 0), 0)
        d = max(int(d_raw or 0), 0)
        soma_produtos[pid]['garrafas'] += g
        soma_produtos[pid]['doses'] += d

    if soma_produtos:
        produtos_map = {str(p.id): p for p in Produto.objects.filter(id__in=soma_produtos.keys())}
        for pid, tot in soma_produtos.items():
            produto = produtos_map.get(str(pid))
            if not produto:
                continue
            EventoProduto.objects.create(
                evento=evento,
                produto=produto,
                garrafas=int(tot['garrafas']),
                doses=int(tot['doses']),
            )

    # ---------- Alimentos ----------
    alimentos_ids = request.POST.getlist('alimento_id[]')
    qts_list = request.POST.getlist('alimento_qtd[]')

    soma_alimentos = defaultdict(lambda: Decimal('0'))
    for aid, q_raw in zip_longest(alimentos_ids, qts_list, fillvalue='0'):
        aid = (aid or '').strip()
        if not aid:
            continue
        try:
            q = Decimal(str(q_raw).replace(',', '.'))
            if q < 0: q = Decimal('0')
        except Exception:
            q = Decimal('0')
        soma_alimentos[aid] += q

    if soma_alimentos:
        alimentos_map = {str(a.id): a for a in Alimento.objects.filter(id__in=soma_alimentos.keys())}
        for aid, qtd in soma_alimentos.items():
            alimento = alimentos_map.get(str(aid))
            if not alimento:
                continue
            EventoAlimento.objects.create(evento=evento, alimento=alimento, quantidade=qtd)

    messages.success(request, "Evento cadastrado.")
    return redirect('pagina_eventos')



@login_required
@transaction.atomic
def editar_evento(request, evento_id):
    evento = get_object_or_404(
        Evento.objects.select_related('restaurante', 'responsavel')
              .prefetch_related('produtos__produto', 'alimentos__alimento'),
        id=evento_id
    )

    if evento.status == 'FINALIZADO':
        messages.info(request, "Evento já finalizado.")
        return redirect('pagina_eventos')

    if request.method == 'POST':
        update_fields = []

        # Horas
        horas_raw = (request.POST.get('horas_evento') or '').strip()
        if horas_raw != '':
            try:
                horas_val = Decimal(horas_raw.replace(',', '.'))
                if horas_val < 0:
                    raise InvalidOperation
                evento.horas = horas_val
                update_fields.append('horas')
            except (InvalidOperation, ValueError):
                pass

        # Número de pessoas
        pessoas_raw = (request.POST.get('numero_pessoas_evento') or '').strip()
        if pessoas_raw != '':
            try:
                pessoas_val = int(pessoas_raw)
                if pessoas_val < 0:
                    raise ValueError
                evento.numero_pessoas = pessoas_val
                update_fields.append('numero_pessoas')
            except ValueError:
                pass

        # (Opcional) trocar restaurante na edição:
        # restaurante_evento_id = (request.POST.get('restaurante_evento') or '').strip()
        # if restaurante_evento_id != '':
        #     try:
        #         evento.restaurante = Restaurante.objects.get(id=int(restaurante_evento_id))
        #         update_fields.append('restaurante')
        #     except (Restaurante.DoesNotExist, ValueError):
        #         pass

        if update_fields:
            evento.save(update_fields=update_fields)

        # Remoções
        del_prod_ids = request.POST.getlist('del_prod[]')
        if del_prod_ids:
            EventoProduto.objects.filter(evento=evento, id__in=del_prod_ids).delete()

        del_ali_ids = request.POST.getlist('del_ali[]')
        if del_ali_ids:
            EventoAlimento.objects.filter(evento=evento, id__in=del_ali_ids).delete()

        # Atualizações
        for ep in evento.produtos.all():
            if not EventoProduto.objects.filter(id=ep.id).exists():
                continue
            g = (request.POST.get(f'prod_g_{ep.id}', '') or '').strip()
            d = (request.POST.get(f'prod_d_{ep.id}', '') or '').strip()
            try: ep.garrafas = max(int(g or 0), 0)
            except: ep.garrafas = 0
            try: ep.doses = max(int(d or 0), 0)
            except: ep.doses = 0
            ep.save()

        for ea in evento.alimentos.all():
            if not EventoAlimento.objects.filter(id=ea.id).exists():
                continue
            q = (request.POST.get(f'ali_q_{ea.id}', '') or '').strip().replace(',', '.')
            try:
                val = Decimal(q or '0')
                ea.quantidade = val if val >= 0 else Decimal('0')
            except (InvalidOperation, ValueError):
                ea.quantidade = Decimal('0')
            ea.save()

        # Adições
        novo_prod = request.POST.get('novo_produto')
        novo_g = request.POST.get('novo_garrafas')
        novo_d = request.POST.get('novo_doses')
        if novo_prod:
            try:
                p = Produto.objects.get(id=novo_prod)
                g = max(int(novo_g or 0), 0)
                d = max(int(novo_d or 0), 0)
                EventoProduto.objects.create(evento=evento, produto=p, garrafas=g, doses=d)
            except Exception:
                pass

        novo_ali = request.POST.get('novo_alimento')
        novo_q = (request.POST.get('novo_qtd') or '').replace(',', '.')
        if novo_ali:
            try:
                a = Alimento.objects.get(id=novo_ali)
                qv = Decimal(novo_q or '0')
                EventoAlimento.objects.create(evento=evento, alimento=a, quantidade=max(qv, Decimal('0')))
            except Exception:
                pass

        # Finalização
        if 'finalizar' in request.POST:
            evento.status = 'FINALIZADO'
            evento.finalizado_em = timezone.now()
            evento.supervisor_finalizou = request.user
            evento.save(update_fields=['status', 'finalizado_em', 'supervisor_finalizou'])
            messages.success(request, "Evento finalizado. Ele agora aparece no consolidado/Excel.")
            return redirect('pagina_eventos')

        messages.success(request, "Alterações salvas (itens removidos/atualizados).")
        return redirect('editar_evento', evento_id=evento.id)

    produtos = Produto.objects.all().order_by('nome')
    alimentos = Alimento.objects.filter(ativo=True).order_by('nome')
    return render(request, 'eventos/editar_evento.html', {
        'evento': evento,
        'produtos': produtos,
        'alimentos': alimentos,
        # 'restaurantes': Restaurante.objects.all().order_by('nome'),  # se quiser trocar na edição
    })










@login_required
def salvar_evento(request):
    if request.method == 'POST':
        nome = request.POST.get('nome_evento')
        pessoas_raw = (request.POST.get('numero_pessoas') or '').strip()
        horas_raw = (request.POST.get('horas') or '').strip()

        # saneamento
        try:
            numero_pessoas = int(pessoas_raw) if pessoas_raw != '' else None
            if numero_pessoas is not None and numero_pessoas < 0:
                numero_pessoas = None
        except ValueError:
            numero_pessoas = None

        try:
            horas = Decimal(horas_raw.replace(',', '.')) if horas_raw != '' else None
            if horas is not None and horas < 0:
                horas = None
        except (InvalidOperation, ValueError):
            horas = None

        evento = Evento.objects.create(
            nome=nome,
            responsavel=request.user,
            numero_pessoas=numero_pessoas,
            horas=horas,
        )

        # ---------- Bebidas ----------
        produtos_ids = request.POST.getlist('produto_id[]')
        garrafas_list = request.POST.getlist('garrafas[]')
        doses_list = request.POST.getlist('doses[]')

        for i in range(len(produtos_ids)):
            try:
                produto = Produto.objects.get(id=produtos_ids[i])
                garrafas = int(garrafas_list[i])
                doses = int(doses_list[i])
                if garrafas > 0 or doses > 0:
                    EventoProduto.objects.create(
                        evento=evento, produto=produto,
                        garrafas=garrafas, doses=doses
                    )
            except (Produto.DoesNotExist, ValueError, IndexError):
                continue

        # ---------- ✅ Alimentos ----------
        alimentos_ids = request.POST.getlist('alimento_id[]')
        qts_list = request.POST.getlist('alimento_qtd[]')

        for i in range(len(alimentos_ids)):
            try:
                alimento = Alimento.objects.get(id=alimentos_ids[i])
                qtd = Decimal((qts_list[i] or '0').replace(',', '.'))
                if qtd > 0:
                    EventoAlimento.objects.create(evento=evento, alimento=alimento, quantidade=qtd)
            except (Alimento.DoesNotExist, InvalidOperation, ValueError, IndexError):
                continue

        return redirect('pagina_eventos')


@login_required
def excluir_evento(request, evento_id):
    evento = get_object_or_404(Evento, id=evento_id)
    if evento.status == 'FINALIZADO':
        messages.error(request, "Não é permitido excluir eventos finalizados.")
        return redirect('pagina_eventos')
    evento.delete()
    messages.success(request, "Evento excluído.")
    return redirect('pagina_eventos')




@login_required
def dashboard(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        # se alguém digitar /dashboard sem bar, manda escolher
        request.session['next_after_select'] = request.path
        return redirect('selecionar-bar')

    bar = get_object_or_404(Bar, id=bar_id)

    # Estoque atual
    estoque_qs = EstoqueBar.objects.filter(bar=bar).select_related('produto')
    estoque_agrupado = [
        {
            'produto': e.produto.nome,
            'garrafas': e.quantidade_garrafas,
            'doses': e.quantidade_doses
        }
        for e in estoque_qs
    ]

    # Top 5 (garrafas + doses/10)
    estoque_top_qs = sorted(
        estoque_qs,
        key=lambda e: e.quantidade_garrafas + (e.quantidade_doses / Decimal('10')),
        reverse=True
    )[:5]
    estoque_labels = [e.produto.nome for e in estoque_top_qs]
    estoque_valores = [
        round(e.quantidade_garrafas + (e.quantidade_doses / Decimal('10')), 2)
        for e in estoque_top_qs
    ]

    # Últimas movimentações
    ultimas_requisicoes = RequisicaoProduto.objects.filter(bar=bar).order_by('-data_solicitacao')[:5]
    ultimas_transferencias = TransferenciaBar.objects.filter(origem=bar).order_by('-data_transferencia')[:5]

    # Ranking requisitados
    ranking_qs = (
        RequisicaoProduto.objects
        .filter(bar=bar)
        .values('produto__nome')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )
    produtos_ranking = [r['produto__nome'] for r in ranking_qs]
    totais_ranking = [r['total'] for r in ranking_qs]

    return render(request, 'core/dashboard.html', {
        'bar': bar,
        'estoque': estoque_agrupado,
        'ultimas_requisicoes': ultimas_requisicoes,
        'ultimas_transferencias': ultimas_transferencias,
        'dias': estoque_labels,
        'saidas': estoque_valores,
        'ranking_produtos': produtos_ranking,
        'ranking_totais': totais_ranking,
    })


    

def _to_int0(v):
    try:
        x = int(str(v).strip())
        return x if x > 0 else 0
    except Exception:
        return 0

@login_required
def pagina_perdas(request):
    """Tela de registro de perdas SEM escolher bar: usa o bar logado."""
    hoje = timezone.localdate()

    bar_id = request.session.get('bar_id')
    if not bar_id:
        messages.error(request, "Selecione um bar para continuar.")
        return redirect('selecionar-bar')

    bar = get_object_or_404(Bar, id=bar_id)

    # produtos ativos (se quiser, pode filtrar só os que existem no estoque do bar)
    produtos = Produto.objects.filter(ativo=True).order_by('nome')

    # perdas do dia APENAS do bar logado
    perdas_hoje = (
        PerdaProduto.objects
        .filter(bar=bar, data_registro__date=hoje)
        .select_related('bar', 'produto', 'usuario')
        .order_by('-data_registro')
    )

    # consolidado do dia (do bar logado)
    consolidado = defaultdict(lambda: {'garrafas': 0, 'doses': 0})
    for p in perdas_hoje:
        key = f"[{getattr(p.produto, 'codigo', '')}] {p.produto.nome}" if getattr(p.produto, 'codigo', None) else p.produto.nome
        consolidado[key]['garrafas'] += int(p.garrafas or 0)
        consolidado[key]['doses']    += int(p.doses or 0)

    return render(request, 'perdas/pagina_perdas.html', {
        'bar': bar,
        'produtos': produtos,
        'perdas_hoje': perdas_hoje,
        'consolidado': dict(consolidado),
        'hoje': hoje,
    })


@login_required
@transaction.atomic
def registrar_perda(request):
    """Registra perda debitando do estoque do BAR DA SESSÃO."""
    if request.method != 'POST':
        return redirect('pagina_perdas')

    bar_id = request.session.get('bar_id')
    if not bar_id:
        messages.error(request, "Selecione um bar para continuar.")
        return redirect('selecionar-bar')

    bar        = get_object_or_404(Bar, id=bar_id)
    produto_id = request.POST.get('produto')
    garrafas   = _to_int0(request.POST.get('garrafas'))
    doses      = _to_int0(request.POST.get('doses'))
    motivo     = (request.POST.get('motivo') or 'OUTRO').upper()
    observacao = (request.POST.get('observacao') or '').strip()

    if not produto_id:
        messages.error(request, "Selecione o produto (bebida).")
        return redirect('pagina_perdas')

    if garrafas == 0 and doses == 0:
        messages.error(request, "Informe pelo menos garrafas ou doses para registrar perda.")
        return redirect('pagina_perdas')

    produto = get_object_or_404(Produto, id=produto_id)

    # captura saldo antes
    estoque, _ = EstoqueBar.objects.get_or_create(
        bar=bar, produto=produto,
        defaults={'quantidade_garrafas': Decimal('0'), 'quantidade_doses': Decimal('0')}
    )
    antes_g = Decimal(estoque.quantidade_garrafas)
    antes_d = Decimal(estoque.quantidade_doses)

    # baixa no estoque (bloqueia se insuficiente)
    ok = EstoqueBar.retirar(
        bar=bar, produto=produto,
        garrafas=Decimal(garrafas), doses=Decimal(doses)
    )
    if not ok:
        messages.error(request, "Estoque insuficiente para registrar a perda.")
        return redirect('pagina_perdas')

    # recarrega para pegar saldo depois
    estoque = EstoqueBar.objects.get(bar=bar, produto=produto)
    depois_g = Decimal(estoque.quantidade_garrafas)
    depois_d = Decimal(estoque.quantidade_doses)

    PerdaProduto.objects.create(
        restaurante=bar.restaurante,
        bar=bar,
        produto=produto,
        garrafas=garrafas,
        doses=doses,
        motivo=motivo if motivo in dict(PerdaProduto.MOTIVOS) else 'OUTRO',
        observacao=observacao,
        usuario=request.user,
        estoque_antes_garrafas=antes_g,
        estoque_antes_doses=antes_d,
        estoque_depois_garrafas=depois_g,
        estoque_depois_doses=depois_d,
    )

    messages.success(request, "Perda registrada e descontada do estoque do seu bar.")
    return redirect('pagina_perdas')

@login_required
@transaction.atomic
def excluir_perda(request, perda_id):
    perda = get_object_or_404(PerdaProduto, id=perda_id)
    # devolve ao estoque o que foi perdido (para correções), somente no mesmo dia
    if perda.data_registro.date() == timezone.localdate():
        EstoqueBar.adicionar(
            bar=perda.bar, produto=perda.produto,
            garrafas=Decimal(perda.garrafas), doses=Decimal(perda.doses)
        )
        perda.delete()
        messages.success(request, "Perda removida e estoque restituído.")
    else:
        messages.error(request, "Só é possível excluir perdas registradas hoje.")
    return redirect('pagina_perdas')


@login_required
@require_POST
def marcar_perda_baixada(request, perda_id):
    perda = get_object_or_404(
        PerdaProduto.objects.select_related('bar', 'produto'),
        id=perda_id
    )
    if perda.baixado:
        messages.info(request, "Esta perda já estava marcada como baixada.")
        return redirect('relatorio_perdas')

    perda.baixado = True
    perda.baixado_em = timezone.now()
    perda.baixado_por = request.user
    perda.baixado_obs = (request.POST.get('obs') or '').strip()[:255]
    perda.save()
    messages.success(request, f"Perda de '{perda.produto.nome}' marcada como baixada.")
    return redirect('relatorio_perdas')


@login_required
@require_POST
def desmarcar_perda_baixada(request, perda_id):
    perda = get_object_or_404(PerdaProduto, id=perda_id)
    if not perda.baixado:
        messages.info(request, "Esta perda não estava marcada como baixada.")
        return redirect('relatorio_perdas')

    perda.baixado = False
    perda.baixado_em = None
    perda.baixado_por = None
    perda.baixado_obs = None
    perda.save()
    messages.success(request, "Marca de baixa removida.")
    return redirect('relatorio_perdas')



# ---------- Helpers de normalização ----------
CATEGORIAS_SET = {"DESTILADO", "CERVEJA", "VINHO", "OUTRO"}
UNIDADES_ALIMENTO = {"un","kg","g","porcao","l","ml"}

def _to_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    # pandas às vezes traz "nan" como string
    return "" if s.lower() == "nan" else s

def _to_int_pos(v, default=0):
    try:
        n = int(float(str(v).replace(",", ".").strip()))
        return n if n >= 0 else default
    except Exception:
        return default

def _to_bool(v, default=True):
    s = _to_str(v).lower()
    if s in {"1","true","sim","yes","y"}:
        return True
    if s in {"0","false","nao","não","no","n"}:
        return False
    return default

def _norm_categoria(v):
    s = _to_str(v).upper()
    return s if s in CATEGORIAS_SET else "OUTRO"

def _norm_unidade_alimento(v):
    s = _to_str(v).lower()
    # aceita sinônimos simples
    mapping = {
        "unid":"un", "und":"un", "unidade":"un",
        "porção":"porcao", "porc":"porcao",
        "litro":"l", "lt":"l",
        "mililitro":"ml"
    }
    s = mapping.get(s, s)
    return s if s in UNIDADES_ALIMENTO else "un"


# ---------- Leitura de arquivo ----------
def _read_table_from_uploaded(file_path):
    """
    Lê .xlsx/.xls via pandas (openpyxl) ou .csv. Retorna (columns, rows)
    onde rows é uma lista de dicts.
    """
    name = file_path.lower()
    with default_storage.open(file_path, "rb") as f:
        data = f.read()

    # tenta Excel
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
    else:
        # tenta CSV com auto-detector
        try:
            df = pd.read_csv(io.BytesIO(data))
        except Exception:
            df = pd.read_csv(io.BytesIO(data), sep=";")

    # normaliza nomes de colunas (mostramos como estão; sem transformar)
    cols = list(df.columns)
    # converte para rows (dict)
    rows = df.where(pd.notnull(df), None).to_dict(orient="records")
    return cols, rows


# ---------- Assistente de Importação ----------
@login_required
def assistente_importacao(request):
    # (opcional) proteger por permissão; se preferir só superuser, troque o if.
    # if not request.user.is_superuser:
    #     messages.error(request, "Acesso restrito.")
    #     return redirect("dashboard")
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='importacao').exists():
        messages.error(request, "Você não tem permissão para acessar a importação.")
        return redirect("dashboard")

    step = request.POST.get("step") or request.GET.get("step") or "1"

    # ---------- PASSO 1: upload ----------
    if step == "1" and request.method in {"GET","POST"}:
        if request.method == "POST" and request.FILES.get("planilha"):
            tipo = request.POST.get("tipo")  # "produtos" ou "alimentos"
            if tipo not in {"produtos","alimentos"}:
                messages.error(request, "Selecione o tipo (Produtos ou Alimentos).")
                return redirect("assistente_importacao")

            up = request.FILES["planilha"]
            # salva temporariamente
            ext = up.name.split(".")[-1].lower()
            saved_name = f"imports/{uuid.uuid4()}.{ext}"
            default_storage.save(saved_name, up)

            # lê para montar mapeamento
            try:
                cols, rows = _read_table_from_uploaded(saved_name)
            except Exception as e:
                default_storage.delete(saved_name)
                messages.error(request, f"Falha lendo a planilha: {e}")
                return redirect("assistente_importacao")

            # preview (até 10)
            preview = rows[:10]

            # colunas sugeridas por heurística
            sugest = {
                "codigo": _suggest_col(cols, ["codigo","código","cod","code","id"]),
                "nome":   _suggest_col(cols, ["produto","nome","descrição","descricao","item"]),
            }
            if tipo == "produtos":
                sugest.update({
                    "categoria": _suggest_col(cols, ["categoria","tipo","classe"]),
                    "doses_por_garrafa": _suggest_col(cols, ["doses","doses_por_garrafa","shots","dose"]),
                    "unidade_medida": _suggest_col(cols, ["unidade","unidade_medida","um"]),
                    "ativo": _suggest_col(cols, ["ativo","status"]),
                })
            else:
                sugest.update({
                    "unidade": _suggest_col(cols, ["unidade","um","medida"]),
                    "ativo": _suggest_col(cols, ["ativo","status"]),
                })

            return render(request, "core/importador.html", {
                "step": 2,
                "file_id": saved_name,
                "tipo": tipo,
                "columns": cols,
                "preview": preview,
                "sugestoes": sugest,
            })

        # GET (primeira visita) ou POST sem arquivo
        return render(request, "core/importador.html", {"step": 1})

    # ---------- PASSO 2: confirmar/importar ----------
    if step == "2" and request.method == "POST":
        file_id = request.POST.get("file_id")
        tipo = request.POST.get("tipo")
        if not file_id or not default_storage.exists(file_id):
            messages.error(request, "Arquivo temporário não encontrado. Refaça o upload.")
            return redirect("assistente_importacao")

        # mapeamento
        col_codigo = request.POST.get("map_codigo") or ""
        col_nome   = request.POST.get("map_nome") or ""
        atualizar  = request.POST.get("atualizar") == "on"

        if not col_codigo or not col_nome:
            messages.error(request, "Mapeie pelo menos Código e Nome.")
            return redirect("assistente_importacao")

        # campos opcionais por tipo
        if tipo == "produtos":
            col_categoria = request.POST.get("map_categoria") or ""
            col_doses = request.POST.get("map_doses") or ""
            col_um = request.POST.get("map_um") or ""
            col_ativo = request.POST.get("map_ativo") or ""
        else:
            col_unid = request.POST.get("map_unidade") or ""
            col_ativo = request.POST.get("map_ativo") or ""

        # lê novamente o arquivo
        try:
            cols, rows = _read_table_from_uploaded(file_id)
        except Exception as e:
            default_storage.delete(file_id)
            messages.error(request, f"Falha lendo a planilha: {e}")
            return redirect("assistente_importacao")

        criados = 0
        atualizados = 0
        pulados = 0
        erros = []

        if tipo == "produtos":
            for i, r in enumerate(rows, start=2):  # start=2 para numerar como Excel (cabeçalho na 1)
                codigo = _to_str(r.get(col_codigo))
                nome = _to_str(r.get(col_nome))
                if not codigo or not nome:
                    pulados += 1
                    erros.append(f"Linha {i}: código/nome ausente.")
                    continue

                categoria = _norm_categoria(r.get(col_categoria)) if col_categoria else "OUTRO"
                doses_pg = _to_int_pos(r.get(col_doses)) if col_doses else None
                um = _to_str(r.get(col_um)) or "un"
                ativo = _to_bool(r.get(col_ativo), default=True) if col_ativo else True

                obj, created = Produto.objects.get_or_create(
                    codigo=codigo,
                    defaults={
                        "nome": nome.title(),
                        "categoria": categoria,
                        "doses_por_garrafa": doses_pg,
                        "unidade_medida": um,
                        "ativo": ativo,
                    }
                )
                if created:
                    criados += 1
                else:
                    if atualizar:
                        obj.nome = nome
                        obj.categoria = categoria
                        obj.doses_por_garrafa = doses_pg
                        obj.unidade_medida = um
                        obj.ativo = ativo
                        obj.save()
                        atualizados += 1
                    else:
                        pulados += 1

        else:  # alimentos
            for i, r in enumerate(rows, start=2):
                codigo = _to_str(r.get(col_codigo))
                nome = _to_str(r.get(col_nome))
                if not codigo or not nome:
                    pulados += 1
                    erros.append(f"Linha {i}: código/nome ausente.")
                    continue

                unidade = _norm_unidade_alimento(r.get(col_unid)) if col_unid else "un"
                ativo = _to_bool(r.get(col_ativo), default=True) if col_ativo else True

                obj, created = Alimento.objects.get_or_create(
                    codigo=codigo,
                    defaults={"nome": nome.title(), "unidade": unidade, "ativo": ativo}
                )
                if created:
                    criados += 1
                else:
                    if atualizar:
                        obj.nome = nome
                        obj.unidade = unidade
                        obj.ativo = ativo
                        obj.save()
                        atualizados += 1
                    else:
                        pulados += 1

        # limpa arquivo temporário
        try:
            default_storage.delete(file_id)
        except Exception:
            pass

        messages.success(request, f"Importação concluída: {criados} criados, {atualizados} atualizados, {pulados} pulados.")
        if erros:
            # mostra só os 15 primeiros para não poluir
            for e in erros[:15]:
                messages.warning(request, e)
            if len(erros) > 15:
                messages.info(request, f"... e mais {len(erros)-15} linhas com problemas.")
        return redirect("assistente_importacao")

    # fallback
    return redirect("assistente_importacao")


def _suggest_col(columns, candidates):
    """
    Tenta sugerir automaticamente a coluna certa baseado em possíveis nomes.
    """
    low = [c.lower() for c in columns]
    for cand in candidates:
        if cand in low:
            idx = low.index(cand)
            return columns[idx]
    # match parcial
    for i, c in enumerate(low):
        for cand in candidates:
            if cand in c:
                return columns[i]
    return ""





#                                                                                         SEÇÃO DE RELATÓRIOS

@login_required
def relatorios_view(request):
    # 🔒 Verificação de permissão
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='relatorios').exists():
        messages.error(request, "Você não tem permissão para acessar a página de relatórios.")
        return redirect('dashboard')

    return render(request, 'core/relatorios.html')




@login_required
def relatorio_saida_estoque(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return render(request, 'erro.html', {'mensagem': 'Nenhum bar selecionado.'})

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    produto_query = request.GET.get('produto', '').strip()

    requisicoes = RequisicaoProduto.objects.filter(
        bar_id=bar_id,
        status__in=['APROVADA', 'NEGADA', 'FALHA_ESTOQUE']
    )

    # Aplicar os filtros de produto antes de cortar
    if produto_query:
        requisicoes = requisicoes.filter(produto__nome__icontains=produto_query)

    if mes and ano:
        requisicoes = requisicoes.filter(data_solicitacao__month=int(mes), data_solicitacao__year=int(ano))
    elif mes:
        requisicoes = requisicoes.filter(data_solicitacao__month=int(mes))
    elif ano:
        requisicoes = requisicoes.filter(data_solicitacao__year=int(ano))

    # Se nenhum filtro de data foi usado, limita a 50 últimas
    if not mes and not ano:
        requisicoes = requisicoes.order_by('-data_solicitacao')[:50]
    else:
        requisicoes = requisicoes.order_by('-data_solicitacao')

    context = {
        'requisicoes': requisicoes,
        'mes': mes or '',
        'ano': ano or '',
        'produto': produto_query or '',
    }

    return render(request, 'core/relatorios/saida_estoque.html', context)






@login_required
def relatorio_consolidado_view(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return render(request, 'erro.html', {'mensagem': 'Nenhum bar selecionado.'})

    # Filtros de mês e ano, padrão para o mês atual
    hoje = timezone.now()
    mes = request.GET.get('mes', f"{hoje.month:02}")
    ano = request.GET.get('ano', str(hoje.year))

    # Buscar requisições aprovadas no período
    requisicoes = RequisicaoProduto.objects.filter(
        bar_id=bar_id,
        status='APROVADA',
        data_solicitacao__month=int(mes),
        data_solicitacao__year=int(ano)
    )

    dados_agrupados = defaultdict(lambda: defaultdict(dict))  # data -> produto_obj -> dados

    for req in requisicoes:
        data_str = req.data_solicitacao.date().strftime('%d/%m/%Y')
        produto = req.produto

        # Acumular quantidade requisitada
        dados_agrupados[data_str][produto]['quantidade_requisitada'] = \
            dados_agrupados[data_str][produto].get('quantidade_requisitada', 0) + float(req.quantidade_solicitada)

        # Buscar contagem do dia
        contagem = ContagemBar.objects.filter(
            bar_id=bar_id,
            produto=produto,
            data_contagem__date=req.data_solicitacao.date()
        ).order_by('-data_contagem').first()

        if contagem:
            dados_agrupados[data_str][produto]['cheias'] = contagem.quantidade_garrafas_cheias
            dados_agrupados[data_str][produto]['doses'] = float(contagem.quantidade_doses_restantes)
        else:
            dados_agrupados[data_str][produto]['cheias'] = 0
            dados_agrupados[data_str][produto]['doses'] = 0.0

        # Cálculo da diferença (apenas em garrafas)
        requisitado = dados_agrupados[data_str][produto]['quantidade_requisitada']
        cheias = dados_agrupados[data_str][produto]['cheias']
        dados_agrupados[data_str][produto]['diferenca'] = abs(cheias - requisitado)


    # Converter Produto -> produto.nome e ordenar datas de forma decrescente
    dados_agrupados_final = dict()
    for data in sorted(dados_agrupados.keys(), reverse=True):
        produtos_formatados = {
            produto_obj.nome: dados
            for produto_obj, dados in dados_agrupados[data].items()
        }
        dados_agrupados_final[data] = produtos_formatados

    context = {
        'dados_agrupados': dados_agrupados_final,
        'mes': mes,
        'ano': ano,
    }

    return render(request, 'core/relatorios/consolidado_diferenca.html', context)


SHIFT_START_HOUR = 19  # início do "dia operacional": 19:00

@login_required
def relatorio_contagem_atual(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return render(request, 'erro.html', {'mensagem': 'Nenhum bar selecionado.'})

    bar_atual = Bar.objects.get(id=bar_id)
    restaurante = bar_atual.restaurante
    bares = Bar.objects.filter(restaurante=restaurante).order_by('nome')

    # --------- Filtro de data ---------
    filtro_data_str = (request.GET.get('data') or '').strip()   # formato: YYYY-MM-DD
    modo = (request.GET.get('modo') or 'operacional').lower()   # 'operacional' | 'calendario'
    use_range = False
    inicio = fim = None

    if filtro_data_str:
        try:
            dia = datetime.strptime(filtro_data_str, '%Y-%m-%d').date()
            tz = timezone.get_current_timezone()

            if modo == 'calendario':
                # 00:00 -> 23:59 do dia escolhido
                inicio = timezone.make_aware(datetime.combine(dia, time(0, 0, 0)), tz)
                fim    = inicio + timedelta(days=1)
            else:
                # Dia operacional: 19:00 do dia escolhido -> 18:59 do dia seguinte
                start_naive = datetime.combine(dia, time(SHIFT_START_HOUR, 0, 0))
                inicio = timezone.make_aware(start_naive, tz)
                fim    = inicio + timedelta(days=1)

            use_range = True
        except ValueError:
            messages.warning(request, "Data inválida no filtro; exibindo contagem atual.")

    dados_por_bar = {}
    somatorio_total = defaultdict(lambda: {'garrafas': 0, 'doses': 0.0, 'produto': None})

    for bar in bares:
        qs = ContagemBar.objects.filter(bar=bar).order_by('-data_contagem')
        if use_range:
            qs = qs.filter(data_contagem__gte=inicio, data_contagem__lt=fim)

        # Pega a ÚLTIMA contagem de cada produto dentro do período (ou no geral, se sem filtro)
        ultima_por_produto = {}
        for c in qs:
            if c.produto_id not in ultima_por_produto:
                ultima_por_produto[c.produto_id] = c

        contagens_finais = list(ultima_por_produto.values())
        dados_por_bar[bar.nome] = contagens_finais

        for c in contagens_finais:
            pid = c.produto_id
            somatorio_total[pid]['produto'] = c.produto
            somatorio_total[pid]['garrafas'] += c.quantidade_garrafas_cheias or 0
            somatorio_total[pid]['doses']    += float(c.quantidade_doses_restantes or 0)

    context = {
        'dados_por_bar': dados_por_bar,
        'restaurante': restaurante,
        'somatorio_total': dict(somatorio_total),
        'filtro_data': filtro_data_str,
        'modo': 'calendario' if modo == 'calendario' else 'operacional',
        'use_range': use_range,
        'inicio_periodo': inicio,
        'fim_periodo': fim,
        'SHIFT_START_HOUR': SHIFT_START_HOUR,
    }
    return render(request, 'core/relatorios/contagem_atual.html', context)



DOSE_ML = Decimal('50')  # ml por dose

@login_required
def relatorio_eventos(request):
    # filtros (mês atual por padrão)
    data_inicio_param = request.GET.get('data_inicio')
    data_fim_param = request.GET.get('data_fim')
    nome_evento = (request.GET.get('nome_evento') or "").strip()
    somente_nao_baixados = request.GET.get('pendentes') == '1'

    # 🔹 NOVO: filtro por restaurante
    restaurante_param = (request.GET.get('restaurante') or '').strip()
    try:
        restaurante_filtro_id = int(restaurante_param) if restaurante_param else None
    except ValueError:
        restaurante_filtro_id = None

    if data_inicio_param and data_fim_param:
        try:
            data_inicio = datetime.strptime(data_inicio_param, "%Y-%m-%d").date()
            data_fim = datetime.strptime(data_fim_param, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            data_inicio = date.today().replace(day=1)
            data_fim = date.today()
    else:
        data_inicio = date.today().replace(day=1)
        data_fim = date.today()

    eventos_qs = (
        Evento.objects
        .filter(data_criacao__date__range=(data_inicio, data_fim))
        .prefetch_related('produtos__produto', 'alimentos__alimento')
        .select_related('responsavel', 'supervisor_finalizou', 'baixado_por', 'restaurante')  # 👈 inclui restaurante
        .order_by('-data_criacao')
    )
    if nome_evento:
        eventos_qs = eventos_qs.filter(nome__icontains=nome_evento)
    if somente_nao_baixados:
        eventos_qs = eventos_qs.filter(baixado_estoque=False, status='FINALIZADO')
    if restaurante_filtro_id:
        eventos_qs = eventos_qs.filter(restaurante_id=restaurante_filtro_id)

    # ✅ dois consolidados: bebidas e alimentos
    consolidado_bebidas = defaultdict(lambda: {'garrafas': 0, 'doses': Decimal('0'), 'ml': Decimal('0')})
    consolidado_alimentos = defaultdict(lambda: {'quantidade': Decimal('0.00'), 'unidade': ''})

    eventos = []
    for ev in eventos_qs:
        # --- bebidas (por evento)
        total_g, total_d, total_ml = 0, Decimal('0'), Decimal('0')
        itens_bebidas = []
        for item in ev.produtos.all():
            g = int(item.garrafas or 0)
            d = Decimal(item.doses or 0)
            ml = d * DOSE_ML

            itens_bebidas.append({'produto': item.produto.nome, 'garrafas': g, 'doses': d, 'ml': ml})
            total_g += g; total_d += d; total_ml += ml

            # consolida
            nome_prod = item.produto.nome
            consolidado_bebidas[nome_prod]['garrafas'] += g
            consolidado_bebidas[nome_prod]['doses'] += d
            consolidado_bebidas[nome_prod]['ml'] += ml

        # --- alimentos (por evento)
        itens_alimentos = []
        total_qtd_alimentos = Decimal('0')
        for ali in ev.alimentos.all():
            nome = ali.alimento.nome
            qtd = Decimal(ali.quantidade or 0)
            uni = ali.alimento.unidade or ''
            itens_alimentos.append({'alimento': nome, 'quantidade': qtd, 'unidade': uni})
            total_qtd_alimentos += qtd

            # consolida
            consolidado_alimentos[nome]['quantidade'] += qtd
            consolidado_alimentos[nome]['unidade'] = uni

        eventos.append({
            'obj': ev,
            'data': localtime(ev.data_criacao),
            'responsavel': getattr(ev, 'responsavel', ''),
            # ✅ novos campos:
            'pessoas': ev.numero_pessoas,
            'horas': ev.horas,
            'restaurante_nome': ev.restaurante.nome if getattr(ev, 'restaurante', None) else None,  # 👈 para exibir facilmente
            # blocos:
            'itens_bebidas': itens_bebidas,
            'totais_bebidas': {'garrafas': total_g, 'doses': total_d, 'ml': total_ml},
            'itens_alimentos': itens_alimentos,
            'total_qtd_alimentos': total_qtd_alimentos,
        })

    # ordenações amigáveis
    consolidado_bebidas = OrderedDict(sorted(consolidado_bebidas.items(), key=lambda kv: kv[0].lower()))
    consolidado_alimentos = OrderedDict(sorted(consolidado_alimentos.items(), key=lambda kv: kv[0].lower()))

    # 🔹 lista para o dropdown
    restaurantes = Restaurante.objects.all().order_by('nome')

    return render(request, 'core/relatorios/relatorio_eventos.html', {
        'eventos': eventos,
        'consolidado_bebidas': consolidado_bebidas,
        'consolidado_alimentos': consolidado_alimentos,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'nome_evento': nome_evento,
        'somente_nao_baixados': '1' if somente_nao_baixados else '',
        'restaurantes': restaurantes,                    # 👈 para o filtro
        'selected_restaurante': restaurante_filtro_id,   # 👈 para marcar selecionado
    })



@login_required
@transaction.atomic
def marcar_evento_baixado(request, evento_id):
    if request.method != 'POST':
        return redirect('relatorio_eventos')

    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='relatorios').exists():
        messages.error(request, "Você não tem permissão para marcar baixa.")
        return redirect('relatorio_eventos')

    # ✅ Lock apenas na tabela do Evento (sem joins)
    try:
        evento = Evento.objects.select_for_update().get(id=evento_id)
    except Evento.DoesNotExist:
        messages.error(request, "Evento não encontrado.")
        return redirect('relatorio_eventos')

    if evento.status != 'FINALIZADO':
        messages.error(request, "Só é possível marcar baixa de eventos FINALIZADOS.")
        return redirect('relatorio_eventos')

    if evento.baixado_estoque:
        who = getattr(evento.baixado_por, 'username', 'alguém') if evento.baixado_por_id else 'alguém'
        when = timezone.localtime(evento.baixado_em).strftime('%d/%m %H:%M') if evento.baixado_em else ''
        messages.info(request, f"Este evento já estava marcado como baixado ({who} em {when}).")
        return redirect('relatorio_eventos')

    obs = (request.POST.get('obs') or '').strip()
    evento.baixado_estoque = True
    evento.baixado_por = request.user
    evento.baixado_em = timezone.now()
    evento.baixado_obs = obs[:255]
    evento.save()

    messages.success(request, "Evento marcado como baixado.")
    return redirect('relatorio_eventos')



@login_required
@transaction.atomic
def desmarcar_evento_baixado(request, evento_id):
    if request.method != 'POST':
        return redirect('relatorio_eventos')

    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='relatorios').exists():
        messages.error(request, "Você não tem permissão para desmarcar baixa.")
        return redirect('relatorio_eventos')

    try:
        evento = Evento.objects.select_for_update().get(id=evento_id)  # ✅ sem select_related
    except Evento.DoesNotExist:
        messages.error(request, "Evento não encontrado.")
        return redirect('relatorio_eventos')

    if not evento.baixado_estoque:
        messages.info(request, "Este evento não estava baixado.")
        return redirect('relatorio_eventos')

    evento.baixado_estoque = False
    evento.baixado_por = None
    evento.baixado_em = None
    # evento.baixado_obs = ""  # se quiser limpar
    evento.save()

    messages.success(request, "Marca de baixa removida.")
    return redirect('relatorio_eventos')





@login_required
def relatorio_diferenca_contagens(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return render(request, 'erro.html', {'mensagem': 'Nenhum bar selecionado.'})

    bar_atual = get_object_or_404(Bar, id=bar_id)
    restaurante = bar_atual.restaurante

    # Todos os bares do restaurante (igual aos outros relatórios)
    bares = Bar.objects.filter(restaurante=restaurante).order_by('nome')

    # Estruturas de saída
    dados_por_bar = {}  # { "Nome do Bar": [ {produto, ultimo, penultimo, diffs...}, ...] }
    somatorio_total = defaultdict(lambda: {
        'produto': None,
        'diff_garrafas': Decimal('0'),
        'diff_doses': Decimal('0'),
    })

    for bar in bares:
        # Pega contagens do bar, mais novas primeiro
        contagens = (
            ContagemBar.objects
            .filter(bar=bar)
            .select_related('produto', 'usuario')
            .order_by('-data_contagem')
        )

        # Para cada produto do bar, vamos guardar as DUAS mais recentes
        # Estrutura: { produto_id: [ultima, penultima] }
        duas_ultimas_por_produto = {}

        for c in contagens:
            pid = c.produto_id
            if pid not in duas_ultimas_por_produto:
                duas_ultimas_por_produto[pid] = [c]  # primeira (última)
            elif len(duas_ultimas_por_produto[pid]) == 1:
                duas_ultimas_por_produto[pid].append(c)  # segunda (penúltima)
            # se já tem 2, ignora

        linhas_bar = []

        for pid, lista in duas_ultimas_por_produto.items():
            ultimo = lista[0]
            penultimo = lista[1] if len(lista) > 1 else None

            # Converte doses para Decimal pra evitar bug de float
            u_g = Decimal(ultimo.quantidade_garrafas_cheias or 0)
            u_d = Decimal(ultimo.quantidade_doses_restantes or 0)

            if penultimo:
                p_g = Decimal(penultimo.quantidade_garrafas_cheias or 0)
                p_d = Decimal(penultimo.quantidade_doses_restantes or 0)

                diff_g = u_g - p_g
                diff_d = u_d - p_d
            else:
                diff_g = None
                diff_d = None

            # Alimenta o somatório consolidado (só se tiver penúltima)
            if diff_g is not None and diff_d is not None:
                somatorio_total[pid]['produto'] = ultimo.produto
                somatorio_total[pid]['diff_garrafas'] += diff_g
                somatorio_total[pid]['diff_doses'] += diff_d

            linhas_bar.append({
                'produto': ultimo.produto,
                'ultimo': ultimo,
                'penultimo': penultimo,
                'u_g': u_g, 'u_d': u_d,
                'p_g': (p_g if penultimo else None),
                'p_d': (p_d if penultimo else None),
                'diff_g': diff_g,
                'diff_d': diff_d,
            })

        # Ordena por nome do produto dentro do bar (fica mais amigável)
        linhas_bar.sort(key=lambda x: x['produto'].nome.lower())

        dados_por_bar[bar.nome] = linhas_bar

    # Ordena somatório total por nome do produto
    somatorio_total_dict = dict(sorted(
        somatorio_total.items(),
        key=lambda kv: kv[1]['produto'].nome.lower() if kv[1]['produto'] else ''
    ))

    context = {
        'restaurante': restaurante,
        'dados_por_bar': dados_por_bar,
        'somatorio_total': somatorio_total_dict,
    }
    return render(request, 'core/relatorios/diferenca_contagens.html', context)


@login_required
def relatorio_perdas(request):
    """
    Relatório de perdas com filtros por bar, produto (nome/código), motivo, período
    e opção de exibir apenas perdas NÃO baixadas (pendentes).
    """
    # 🔒 mesma permissão do hub de relatórios
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='relatorios').exists():
        messages.error(request, "Você não tem permissão para acessar a página de relatórios.")
        return redirect('dashboard')

    restaurante_id = request.session.get('restaurante_id')
    if not restaurante_id:
        messages.error(request, "Restaurante não selecionado.")
        return redirect('dashboard')

    # Bares do restaurante para o filtro
    bares = Bar.objects.filter(restaurante_id=restaurante_id).order_by('nome')

    # ------- filtros -------
    bar_id    = (request.GET.get('bar') or '').strip()
    produto_q = (request.GET.get('produto') or '').strip()   # nome ou código
    motivo    = (request.GET.get('motivo') or '').strip()    # value de PerdaProduto.MOTIVOS ou vazio
    somente_nao_baixados = (request.GET.get('pendentes') in ('1', 'true', 'on'))

    # Período: default = mês atual
    hoje = timezone.localdate()
    data_inicio_str = request.GET.get('inicio')
    data_fim_str    = request.GET.get('fim')

    if data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
            data_fim    = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
        except ValueError:
            data_inicio = hoje.replace(day=1)
            data_fim    = hoje
    else:
        data_inicio = hoje.replace(day=1)
        data_fim    = hoje

    # ------- base query -------
    qs = (
        PerdaProduto.objects
        .filter(
            restaurante_id=restaurante_id,
            data_registro__date__range=(data_inicio, data_fim)
        )
        .select_related('bar', 'produto', 'usuario')
    )

    if bar_id:
        qs = qs.filter(bar_id=bar_id)

    if produto_q:
        qs = qs.filter(
            Q(produto__nome__icontains=produto_q) |
            Q(produto__codigo__icontains=produto_q)
        )

    if motivo:
        qs = qs.filter(motivo=motivo)

    if somente_nao_baixados:
        qs = qs.filter(baixado=False)

    # ordenar mais recentes primeiro
    qs = qs.order_by('-data_registro')

    # ------- totais gerais -------
    agreg = qs.aggregate(
        total_garrafas=Sum('garrafas'),
        total_doses=Sum('doses')
    )
    total_garrafas = agreg['total_garrafas'] or 0
    total_doses    = agreg['total_doses'] or 0

    # ------- consolidados -------
    # por produto
    por_produto = (
        qs.values('produto__id', 'produto__nome', 'produto__codigo')
          .annotate(garrafas=Sum('garrafas'), doses=Sum('doses'))
          .order_by('produto__nome')
    )

    # por bar
    por_bar = (
        qs.values('bar__id', 'bar__nome')
          .annotate(garrafas=Sum('garrafas'), doses=Sum('doses'))
          .order_by('bar__nome')
    )

    # produtos para datalist (auto-complete no filtro)
    produtos_lista = Produto.objects.filter(ativo=True).order_by('nome')

    context = {
        'bares': bares,
        'itens': qs,  # detalhe
        'por_produto': por_produto,
        'por_bar': por_bar,
        'total_garrafas': total_garrafas,
        'total_doses': total_doses,

        # filtros (eco no form)
        'bar_id': bar_id,
        'produto_q': produto_q,
        'motivo': motivo,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'somente_nao_baixados': somente_nao_baixados,

        'produtos_lista': produtos_lista,
        'MOTIVOS': getattr(PerdaProduto, 'MOTIVOS', ()),
    }
    return render(request, 'core/relatorios/perdas.html', context)





#                                                                                     SEÇÃO DE EXPORTAÇÃO DE EXPORTAÇÃO EXCEL


# ===================== Helpers =====================

DOSE_ML = Decimal("50")  # ml por dose

def auto_fit(ws, min_w=10, max_w=45):
    """Ajusta a largura das colunas pelo maior conteúdo renderizado."""
    for col in ws.columns:
        length = 0
        idx = col[0].column
        for cell in col:
            s = '' if cell.value is None else str(cell.value)
            length = max(length, len(s))
        ws.column_dimensions[get_column_letter(idx)].width = max(min_w, min(max_w, length + 2))

def style_header_row(ws, row_idx, fill_color="F1F5FF"):
    """Aplica estilo de cabeçalho na linha indicada (ws[row_idx])."""
    fill = PatternFill("solid", fgColor=fill_color)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for c in ws[row_idx]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        c.border = border
    ws.row_dimensions[row_idx].height = 20

def style_body_borders(ws, r1, r2, c1, c2):
    """Borda fininha no corpo da tabela (inclusive totais)."""
    border = Border(
        left=Side(style="thin", color="EEEEEE"),
        right=Side(style="thin", color="EEEEEE"),
        top=Side(style="thin", color="EEEEEE"),
        bottom=Side(style="thin", color="EEEEEE"),
    )
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


# ===================== View =====================

@login_required
def exportar_relatorio_eventos_excel(request):
    data_inicio_str = request.GET.get('data_inicio') or ""
    data_fim_str = request.GET.get('data_fim') or ""
    nome_evento = (request.GET.get('nome_evento') or "").strip()

    # 🔹 filtro por restaurante
    restaurante_param = (request.GET.get('restaurante') or '').strip()
    try:
        restaurante_filtro_id = int(restaurante_param) if restaurante_param else None
    except ValueError:
        restaurante_filtro_id = None

    # Período padrão (mês atual)
    if data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
            data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            data_inicio = date.today().replace(day=1)
            data_fim = date.today()
    else:
        data_inicio = date.today().replace(day=1)
        data_fim = date.today()

    eventos_qs = (
        Evento.objects
        .filter(data_criacao__date__range=(data_inicio, data_fim))
        .prefetch_related('produtos__produto', 'alimentos__alimento')
        .select_related('responsavel', 'restaurante')
        .order_by('-data_criacao')
    )
    if nome_evento:
        eventos_qs = eventos_qs.filter(nome__icontains=nome_evento)
    if restaurante_filtro_id:
        eventos_qs = eventos_qs.filter(restaurante_id=restaurante_filtro_id)

    # Texto do filtro
    filtro_txt = f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    if nome_evento:
        filtro_txt += f" | Evento contém: {nome_evento}"
    if restaurante_filtro_id:
        try:
            rnome = Restaurante.objects.only('nome').get(id=restaurante_filtro_id).nome
            filtro_txt += f" | Restaurante: {rnome}"
        except Restaurante.DoesNotExist:
            filtro_txt += " | Restaurante: (inválido)"

    # Consolidações
    consolidado_bebidas = defaultdict(lambda: {'garrafas': 0, 'doses': Decimal('0'), 'ml': Decimal('0')})
    consolidado_alimentos = defaultdict(lambda: {'quantidade': Decimal('0.00'), 'unidade': ''})

    # Linhas detalhadas
    detalhado = []  # Evento, Restaurante, Data, Pessoas, Horas, Tipo, Item, Garrafas, Doses, ML, Quantidade, Unidade

    # Agrupado por evento (para as abas "Por Evento" e "Eventos (lista)")
    eventos_group = []

    for ev in eventos_qs:
        total_g = 0
        total_d = Decimal('0')
        total_ml = Decimal('0')
        total_qtd_alimentos_evento = Decimal('0')
        itens_ev_bebidas = []
        itens_ev_alimentos = []
        rnome = ev.restaurante.nome if getattr(ev, 'restaurante', None) else None

        # ---- Bebidas
        for item in ev.produtos.all():
            prod = item.produto.nome
            gar = int(item.garrafas or 0)
            dos = Decimal(item.doses or 0)
            ml = dos * DOSE_ML

            consolidado_bebidas[prod]['garrafas'] += gar
            consolidado_bebidas[prod]['doses'] += dos
            consolidado_bebidas[prod]['ml'] += ml

            detalhado.append({
                'evento': ev.nome,
                'restaurante': rnome,
                'data': localtime(ev.data_criacao),
                'pessoas': ev.numero_pessoas,
                'horas': ev.horas,
                'tipo': 'Bebida',
                'item': prod,
                'garrafas': gar,
                'doses': dos,
                'ml': ml,
                'quantidade': None,
                'unidade': None,
            })

            itens_ev_bebidas.append({'produto': prod, 'garrafas': gar, 'doses': dos, 'ml': ml})
            total_g += gar
            total_d += dos
            total_ml += ml

        # ---- Alimentos
        for ali in ev.alimentos.all():
            nome = ali.alimento.nome
            qtd = Decimal(ali.quantidade or 0)
            uni = ali.alimento.unidade or ''

            consolidado_alimentos[nome]['quantidade'] += qtd
            consolidado_alimentos[nome]['unidade'] = uni

            detalhado.append({
                'evento': ev.nome,
                'restaurante': rnome,
                'data': localtime(ev.data_criacao),
                'pessoas': ev.numero_pessoas,
                'horas': ev.horas,
                'tipo': 'Alimento',
                'item': nome,
                'garrafas': None,
                'doses': None,
                'ml': None,
                'quantidade': qtd,
                'unidade': uni,
            })

            itens_ev_alimentos.append({'alimento': nome, 'quantidade': qtd, 'unidade': uni})
            total_qtd_alimentos_evento += qtd

        eventos_group.append({
            'nome': ev.nome,
            'data': localtime(ev.data_criacao),
            'responsavel': getattr(ev, 'responsavel', ''),
            'pessoas': ev.numero_pessoas,
            'horas': ev.horas,
            'restaurante_nome': rnome,
            'status': ev.status,
            'baixado': ev.baixado_estoque,
            'itens_bebidas': itens_ev_bebidas,
            'totais_bebidas': {'garrafas': total_g, 'doses': total_d, 'ml': total_ml},
            'itens_alimentos': itens_ev_alimentos,
            'total_alimentos_qtd': total_qtd_alimentos_evento,  # 👈 para a aba de lista
        })

    # === Workbook
    wb = Workbook()

    # ===================== Aba 1: Consolidado Bebidas =====================
    ws1 = wb.active
    ws1.title = "Consolidado Bebidas"

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws1.cell(row=1, column=1, value="Relatório de Eventos — Consolidado de Bebidas").font = Font(bold=True, size=14)

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws1.cell(row=2, column=1, value=filtro_txt).font = Font(italic=True, size=11)

    ws1.append([])
    ws1.append(["Produto", "Garrafas", "Doses", "Doses (ML)"])
    style_header_row(ws1, 4)

    r = 5
    tot_g, tot_d, tot_ml = 0, Decimal('0'), Decimal('0')
    for prod in sorted(consolidado_bebidas.keys(), key=lambda s: s.lower()):
        d = consolidado_bebidas[prod]
        ws1.cell(row=r, column=1, value=prod)
        ws1.cell(row=r, column=2, value=int(d['garrafas'])).number_format = "0"
        ws1.cell(row=r, column=3, value=float(d['doses'])).number_format = "0.00"
        ws1.cell(row=r, column=4, value=float(d['ml'])).number_format = "0.00"
        tot_g += int(d['garrafas']); tot_d += d['doses']; tot_ml += d['ml']
        r += 1

    ws1.cell(row=r, column=1, value="Total").font = Font(bold=True)
    ws1.cell(row=r, column=2, value=tot_g).number_format = "0"
    ws1.cell(row=r, column=3, value=float(tot_d)).number_format = "0.00"
    ws1.cell(row=r, column=4, value=float(tot_ml)).number_format = "0.00"

    ws1.freeze_panes = "A5"
    style_body_borders(ws1, 5, r, 1, 4)
    auto_fit(ws1)

    # ===================== Aba 2: Consolidado Alimentos =====================
    wsA = wb.create_sheet("Consolidado Alimentos")

    wsA.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    wsA.cell(row=1, column=1, value="Relatório de Eventos — Consolidado de Alimentos").font = Font(bold=True, size=14)

    wsA.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    wsA.cell(row=2, column=1, value=filtro_txt).font = Font(italic=True, size=11)

    wsA.append([])
    wsA.append(["Alimento", "Quantidade", "Unidade"])
    style_header_row(wsA, 4)

    rA = 5
    for ali in sorted(consolidado_alimentos.keys(), key=lambda s: s.lower()):
        d = consolidado_alimentos[ali]
        wsA.cell(row=rA, column=1, value=ali)
        wsA.cell(row=rA, column=2, value=float(d['quantidade'])).number_format = "0.00"
        wsA.cell(row=rA, column=3, value=d['unidade'])
        rA += 1

    wsA.freeze_panes = "A5"
    style_body_borders(wsA, 5, rA - 1, 1, 3)
    auto_fit(wsA)

    # ===================== Aba 3: Detalhado =====================
    ws2 = wb.create_sheet("Detalhado")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws2.cell(row=1, column=1, value="Relatório de Eventos — Detalhado (Bebidas e Alimentos)").font = Font(bold=True, size=14)

    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws2.cell(row=2, column=1, value=filtro_txt).font = Font(italic=True, size=11)

    ws2.append([])
    ws2.append([
        "Evento", "Restaurante", "Data", "Pessoas", "Horas", "Tipo", "Item",
        "Garrafas", "Doses", "Doses (ML)", "Quantidade", "Unidade"
    ])
    style_header_row(ws2, 4)

    r2 = 5
    for linha in detalhado:
        ws2.cell(row=r2, column=1, value=linha['evento'])
        ws2.cell(row=r2, column=2, value=linha['restaurante'] or "-")
        ws2.cell(row=r2, column=3, value=linha['data'].strftime("%d/%m/%Y %H:%M"))

        if linha['pessoas'] is not None:
            ws2.cell(row=r2, column=4, value=int(linha['pessoas'])).number_format = "0"
        if linha['horas'] is not None:
            try:
                ws2.cell(row=r2, column=5, value=float(linha['horas'])).number_format = "0.00"
            except Exception:
                pass

        ws2.cell(row=r2, column=6, value=linha['tipo'])
        ws2.cell(row=r2, column=7, value=linha['item'])

        if linha['garrafas'] is not None:
            ws2.cell(row=r2, column=8, value=int(linha['garrafas'])).number_format = "0"
        if linha['doses'] is not None:
            ws2.cell(row=r2, column=9, value=float(linha['doses'])).number_format = "0.00"
        if linha['ml'] is not None:
            ws2.cell(row=r2, column=10, value=float(linha['ml'])).number_format = "0.00"

        if linha['quantidade'] is not None:
            ws2.cell(row=r2, column=11, value=float(linha['quantidade'])).number_format = "0.00"
        if linha['unidade'] is not None:
            ws2.cell(row=r2, column=12, value=linha['unidade'])

        r2 += 1

    ws2.freeze_panes = "A5"
    style_body_borders(ws2, 5, r2 - 1, 1, 12)
    auto_fit(ws2)

    # ===================== Aba 4: Por Evento =====================
    ws3 = wb.create_sheet("Por Evento")

    col_count = 8
    current_row = 1

    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
    ws3.cell(row=current_row, column=1, value="Relatório de Eventos — Por Evento").font = Font(bold=True, size=14)
    current_row += 1
    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
    ws3.cell(row=current_row, column=1, value=filtro_txt).font = Font(italic=True, size=11)
    current_row += 2

    for ev in eventos_group:
        ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
        cab = f"Evento: {ev['nome']}  |  Data: {ev['data'].strftime('%d/%m/%Y %H:%M')}"
        if ev['responsavel']:
            cab += f"  |  Resp.: {ev['responsavel']}"
        if ev.get('pessoas') is not None:
            cab += f"  |  Pessoas: {ev['pessoas']}"
        if ev.get('horas') is not None:
            try:
                cab += f"  |  Horas: {float(ev['horas']):.2f}"
            except Exception:
                cab += "  |  Horas: -"
        if ev.get('restaurante_nome'):
            cab += f"  |  Restaurante: {ev['restaurante_nome']}"
        ws3.cell(row=current_row, column=1, value=cab).font = Font(bold=True, size=12)
        current_row += 1

        ws3.append([]); current_row += 1
        ws3.append(["🍹 Bebidas", "", "", "", "", "", "", ""]); style_header_row(ws3, current_row); current_row += 1
        ws3.append(["Produto", "Garrafas", "Doses", "Doses (ML)", "", "", "", ""]); style_header_row(ws3, current_row)
        first_data_beb = current_row + 1

        for it in ev['itens_bebidas']:
            current_row += 1
            ws3.cell(row=current_row, column=1, value=it['produto'])
            ws3.cell(row=current_row, column=2, value=int(it['garrafas'])).number_format = "0"
            ws3.cell(row=current_row, column=3, value=float(it['doses'])).number_format = "0.00"
            ws3.cell(row=current_row, column=4, value=float(it['ml'])).number_format = "0.00"

        current_row += 1
        ws3.cell(row=current_row, column=1, value="Subtotal (bebidas)").font = Font(bold=True)
        ws3.cell(row=current_row, column=2, value=int(ev['totais_bebidas']['garrafas'])).number_format = "0"
        ws3.cell(row=current_row, column=3, value=float(ev['totais_bebidas']['doses'])).number_format = "0.00"
        ws3.cell(row=current_row, column=4, value=float(ev['totais_bebidas']['ml'])).number_format = "0.00"
        style_body_borders(ws3, first_data_beb, current_row, 1, 4)

        current_row += 2
        ws3.append([]); style_header_row(ws3, current_row); current_row += 1
        ws3.append(["Alimento", "Quantidade", "Unidade", "", "", "", "", ""]); style_header_row(ws3, current_row)
        first_data_ali = current_row + 1

        if ev['itens_alimentos']:
            for it in ev['itens_alimentos']:
                current_row += 1
                ws3.cell(row=current_row, column=1, value=it['alimento'])
                ws3.cell(row=current_row, column=2, value=float(it['quantidade'])).number_format = "0.00"
                ws3.cell(row=current_row, column=3, value=it['unidade'])
            style_body_borders(ws3, first_data_ali, current_row, 1, 3)
        else:
            current_row += 1
            ws3.cell(row=current_row, column=1, value="Sem lançamentos de alimentos.").font = Font(italic=True)

        current_row += 2

    auto_fit(ws3)

    # ===================== Aba 5: Eventos (lista) =====================
    wsL = wb.create_sheet("Eventos (lista)")

    wsL.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    wsL.cell(row=1, column=1, value="Relatório de Eventos — Lista").font = Font(bold=True, size=14)

    wsL.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    wsL.cell(row=2, column=1, value=filtro_txt).font = Font(italic=True, size=11)

    wsL.append([])
    wsL.append([
        "Evento", "Restaurante", "Data", "Status", "Baixado?",
        "Responsável", "Pessoas", "Horas",
        "Garrafas (tot)", "Doses (tot)", "ML (tot)", "Alimentos (qtd tot)"
    ])
    style_header_row(wsL, 4)

    rL = 5
    for ev in eventos_group:
        wsL.cell(row=rL, column=1, value=ev['nome'])
        wsL.cell(row=rL, column=2, value=ev['restaurante_nome'] or "-")
        wsL.cell(row=rL, column=3, value=ev['data'].strftime("%d/%m/%Y %H:%M"))
        wsL.cell(row=rL, column=4, value=ev['status'])
        wsL.cell(row=rL, column=5, value="Sim" if ev['baixado'] else "Não")
        wsL.cell(row=rL, column=6, value=str(ev['responsavel']) if ev['responsavel'] else "-")

        if ev.get('pessoas') is not None:
            wsL.cell(row=rL, column=7, value=int(ev['pessoas'])).number_format = "0"
        if ev.get('horas') is not None:
            try:
                wsL.cell(row=rL, column=8, value=float(ev['horas'])).number_format = "0.00"
            except Exception:
                pass

        wsL.cell(row=rL, column=9,  value=int(ev['totais_bebidas']['garrafas'])).number_format = "0"
        wsL.cell(row=rL, column=10, value=float(ev['totais_bebidas']['doses'])).number_format = "0.00"
        wsL.cell(row=rL, column=11, value=float(ev['totais_bebidas']['ml'])).number_format = "0.00"
        wsL.cell(row=rL, column=12, value=float(ev['total_alimentos_qtd'])).number_format = "0.00"

        rL += 1

    # Linha de total de eventos
    wsL.append([])
    wsL.cell(row=rL + 1, column=1, value="Total de eventos").font = Font(bold=True)
    wsL.cell(row=rL + 1, column=2, value=len(eventos_group)).number_format = "0"

    wsL.freeze_panes = "A5"
    style_body_borders(wsL, 5, rL - 1, 1, 12)
    auto_fit(wsL)

    # === Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"relatorio_eventos_{slugify(data_inicio)}_a_{slugify(data_fim)}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp






DOSE_PADRAO_ML = 50  # ml por dose

def _auto_fit(ws, min_width=10, extra=2):
    """Ajusta largura das colunas pelo maior conteúdo renderizado."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None:
                continue
            ln = len(str(val))
            if ln > max_len:
                max_len = ln
        width = max(min_width, max_len + extra)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _style_header(ws, row=1, fill_color="228B22"):  # verde floresta
    """Estilo para cabeçalho: fundo, negrito, centralizado e borda."""
    header_fill = PatternFill("solid", fgColor=fill_color)
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[row].height = 20

def _style_table(ws, start_row=2):
    """Borda leve nas linhas de dados + alinhamento."""
    border = Border(
        left=Side(style="thin", color="EEEEEE"),
        right=Side(style="thin", color="EEEEEE"),
        top=Side(style="thin", color="EEEEEE"),
        bottom=Side(style="thin", color="EEEEEE"),
    )
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

@login_required
def exportar_consolidado_eventos_excel(request):
    hoje = timezone.localdate()

    # Apenas eventos FINALIZADOS hoje (espelha a página)
    eventos = (
        Evento.objects.filter(status='FINALIZADO', finalizado_em__date=hoje)
        .prefetch_related('produtos__produto', 'alimentos__alimento')
        .select_related('responsavel', 'supervisor_finalizou')
        .order_by('finalizado_em', 'nome')
    )

    # ---------- Consolidados ----------
    # Bebidas por produto (sem doses equivalentes)
    bebidas = defaultdict(lambda: {'garrafas': 0, 'doses': 0, 'ml': 0})
    # Alimentos por item
    alimentos = defaultdict(lambda: {'quantidade': Decimal('0.00'), 'unidade': ''})

    total_pessoas = 0
    total_eventos = 0

    for ev in eventos:
        total_eventos += 1
        if ev.numero_pessoas:
            total_pessoas += ev.numero_pessoas

        # Bebidas
        for ep in ev.produtos.all():
            p = ep.produto
            key = f"[{getattr(p, 'codigo', '')}] {p.nome}" if getattr(p, "codigo", None) else p.nome

            garrafas_add = int(ep.garrafas or 0)
            doses_add = int(ep.doses or 0)
            ml_add = doses_add * DOSE_PADRAO_ML  # apenas doses avulsas

            bebidas[key]['garrafas'] += garrafas_add
            bebidas[key]['doses'] += doses_add
            bebidas[key]['ml'] += ml_add

        # Alimentos
        for ea in ev.alimentos.all():
            a = ea.alimento
            key = f"[{a.codigo}] {a.nome}"
            alimentos[key]['quantidade'] += (ea.quantidade or Decimal('0'))
            if not alimentos[key]['unidade']:
                alimentos[key]['unidade'] = a.get_unidade_display()

    # ---------- Excel ----------
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_bebidas = wb.create_sheet("Bebidas")
    ws_alimentos = wb.create_sheet("Alimentos")
    ws_eventos = wb.create_sheet("Eventos")

    # ===== Resumo =====
    ws = ws_resumo
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Consolidado de Eventos — {hoje.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["Eventos finalizados", total_eventos, "", ""])
    ws.append(["Total de pessoas (informado)", total_pessoas, "", ""])

    # KPIs de bebidas (sem doses equivalentes)
    total_garrafas = sum(d['garrafas'] for d in bebidas.values())
    total_doses = sum(d['doses'] for d in bebidas.values())
    total_ml = sum(d['ml'] for d in bebidas.values())

    ws.append(["Bebidas - Garrafas", total_garrafas, "", ""])
    ws.append(["Bebidas - Doses (avulsas)", total_doses, "", ""])
    ws.append([f"Bebidas - Volume total (ml @ {DOSE_PADRAO_ML}ml)", total_ml, "", ""])

    # Estilo simples nos KPIs
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(horizontal="right")

    _auto_fit(ws)

    # ===== Bebidas =====
    ws = ws_bebidas
    ws.append(["Produto", "Garrafas", "Doses", f"Volume (ml @ {DOSE_PADRAO_ML}ml)"])
    _style_header(ws, 1)

    for produto, d in sorted(bebidas.items()):
        ws.append([produto, d['garrafas'], d['doses'], d['ml']])

    # Totais
    if ws.max_row > 1:
        ws.append(["TOTAL", total_garrafas, total_doses, total_ml])
        last = ws.max_row
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=last, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F0F0F0")

    _style_table(ws, start_row=2)
    _auto_fit(ws)

    # ===== Alimentos =====
    ws = ws_alimentos
    ws.append(["Alimento", "Quantidade", "Unidade"])
    _style_header(ws, 1)

    for nome, d in sorted(alimentos.items()):
        q = float(d['quantidade'])
        ws.append([nome, q, d['unidade']])

    _style_table(ws, start_row=2)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0.00'
    _auto_fit(ws)

    # ===== Eventos (detalhe) =====
    ws = ws_eventos
    ws.append(["Nome", "Pessoas", "Horas", "Responsável", "Finalizado por", "Finalizado em"])
    _style_header(ws, 1)

    for ev in eventos:
        ws.append([
            ev.nome,
            ev.numero_pessoas if ev.numero_pessoas is not None else "",
            float(ev.horas) if ev.horas is not None else "",
            getattr(ev.responsavel, 'username', '') or '',
            getattr(ev.supervisor_finalizou, 'username', '') or '',
            timezone.localtime(ev.finalizado_em).strftime("%H:%M") if ev.finalizado_em else ""
        ])

    _style_table(ws, start_row=2)
    for cell in ws['B'][1:]:
        cell.number_format = '0'
    for cell in ws['C'][1:]:
        cell.number_format = '0.0'
    _auto_fit(ws)

    # ---------- Resposta ----------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"consolidado_eventos_{hoje:%Y-%m-%d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




@login_required
def exportar_saida_estoque_excel(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return render(request, 'erro.html', {'mensagem': 'Nenhum bar selecionado.'})

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    produto_nome = request.GET.get('produto', '').strip().lower()

    requisicoes = RequisicaoProduto.objects.filter(
        bar_id=bar_id,
        status__in=['APROVADA', 'NEGADA', 'FALHA_ESTOQUE']
    )

    # Aplicar filtros
    if mes and ano:
        requisicoes = requisicoes.filter(data_solicitacao__month=int(mes), data_solicitacao__year=int(ano))
    elif mes:
        requisicoes = requisicoes.filter(data_solicitacao__month=int(mes))
    elif ano:
        requisicoes = requisicoes.filter(data_solicitacao__year=int(ano))

    if produto_nome:
        requisicoes = requisicoes.filter(produto__nome__icontains=produto_nome)

    # Ordenar por data (mais recentes primeiro)
    requisicoes = requisicoes.order_by('-data_solicitacao')

    # Criar planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Saída Estoque Bar"

    headers = ['Produto', 'Quantidade', 'Status', 'Data', 'Solicitado por', 'Aprovado/Negado por']
    ws.append(headers)

    total_quantidade = 0

    for req in requisicoes:
        quantidade = float(req.quantidade_solicitada or 0)
        total_quantidade += quantidade

        ws.append([
            req.produto.nome,
            f"{quantidade:.2f}",
            req.get_status_display(),
            req.data_solicitacao.strftime('%d/%m/%Y %H:%M'),
            req.usuario.username if req.usuario else '-',
            req.usuario_aprovador.username if req.usuario_aprovador else '-',
        ])

    # Linha extra de separação
    ws.append([])

    # Linha de total
    ws.append(['TOTAL', f"{total_quantidade:.2f}", '', '', '', ''])

    # Auto ajuste de colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Gerar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=relatorio_saida_estoque.xlsx'
    wb.save(response)
    return response



@login_required
def relatorio_consolidado_excel_view(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return HttpResponse("Nenhum bar selecionado.", status=400)

    hoje = timezone.now()
    mes = request.GET.get('mes', f"{hoje.month:02}")
    ano = request.GET.get('ano', str(hoje.year))

    requisicoes = RequisicaoProduto.objects.filter(
        bar_id=bar_id,
        status='APROVADA',
        data_solicitacao__month=int(mes),
        data_solicitacao__year=int(ano)
    )

    dados_agrupados = defaultdict(lambda: defaultdict(dict))

    for req in requisicoes:
        data_str = req.data_solicitacao.date().strftime('%d/%m/%Y')
        produto = req.produto

        dados_agrupados[data_str][produto]['quantidade_requisitada'] = \
            dados_agrupados[data_str][produto].get('quantidade_requisitada', 0) + float(req.quantidade_solicitada)

        contagem = ContagemBar.objects.filter(
            bar_id=bar_id,
            produto=produto,
            data_contagem__date=req.data_solicitacao.date()
        ).order_by('-data_contagem').first()

        if contagem:
            dados_agrupados[data_str][produto]['cheias'] = contagem.quantidade_garrafas_cheias
            dados_agrupados[data_str][produto]['doses'] = float(contagem.quantidade_doses_restantes)
        else:
            dados_agrupados[data_str][produto]['cheias'] = 0
            dados_agrupados[data_str][produto]['doses'] = 0.0

        requisitado = dados_agrupados[data_str][produto]['quantidade_requisitada']
        cheias = dados_agrupados[data_str][produto]['cheias']
        dados_agrupados[data_str][produto]['diferenca'] = abs(cheias - requisitado)

    # Criar o Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado Diferença"

    ws.append(["Data", "Produto", "Requisitado", "Contagem (Garrafas)", "Contagem (Doses)", "Diferença"])

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    for data in sorted(dados_agrupados.keys(), reverse=True):
        for produto, dados in dados_agrupados[data].items():
            ws.append([
                data,
                produto.nome,
                f"{dados['quantidade_requisitada']:.2f}",
                dados['cheias'],
                f"{dados['doses']:.2f}",
                f"{dados['diferenca']:.2f}"
            ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_consolidado_{mes}_{ano}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response




@login_required
def exportar_contagem_atual_excel(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return HttpResponse("Nenhum bar selecionado.", status=400)

    bar_atual = Bar.objects.only('id', 'restaurante').get(id=bar_id)
    restaurante = bar_atual.restaurante

    # --------- Filtros vindos do form (iguais à view) ---------
    filtro_data_str = (request.GET.get('data') or '').strip()   # YYYY-MM-DD
    modo = (request.GET.get('modo') or 'operacional').lower()   # 'operacional' | 'calendario'

    use_range = False
    inicio = fim = None
    tz = timezone.get_current_timezone()

    if filtro_data_str:
        try:
            dia = datetime.strptime(filtro_data_str, '%Y-%m-%d').date()
            if modo == 'calendario':
                inicio = timezone.make_aware(datetime.combine(dia, time(0, 0, 0)), tz)
                fim    = inicio + timedelta(days=1)
            else:
                start_naive = datetime.combine(dia, time(SHIFT_START_HOUR, 0, 0))
                inicio = timezone.make_aware(start_naive, tz)
                fim    = inicio + timedelta(days=1)
            use_range = True
        except ValueError:
            # se a data vier inválida, segue como "sem filtro"
            pass

    # --------- Montagem dos dados (mesmo critério da tela) ---------
    bares = Bar.objects.filter(restaurante=restaurante).only('id', 'nome').order_by('nome')

    dados_por_bar = {}
    somatorio_total = defaultdict(lambda: {'garrafas': 0, 'doses': 0.0, 'produto': None})

    for bar in bares:
        qs = ContagemBar.objects.filter(bar=bar).order_by('-data_contagem', '-id')
        if use_range:
            qs = qs.filter(data_contagem__gte=inicio, data_contagem__lt=fim)

        ultima_contagem_por_produto = {}
        for contagem in qs:
            if contagem.produto_id not in ultima_contagem_por_produto:
                ultima_contagem_por_produto[contagem.produto_id] = contagem

        contagens_finais = list(ultima_contagem_por_produto.values())
        dados_por_bar[bar.nome] = contagens_finais

        for c in contagens_finais:
            pid = c.produto_id
            somatorio_total[pid]['produto'] = c.produto
            somatorio_total[pid]['garrafas'] += c.quantidade_garrafas_cheias or 0
            somatorio_total[pid]['doses']    += float(c.quantidade_doses_restantes or 0)

    # --------- Excel ---------
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = workbook.add_worksheet('Contagem Atual')

    # Formatos
    f_title = workbook.add_format({'bold': True, 'font_size': 14})
    f_sub   = workbook.add_format({'italic': True, 'font_color': '#555555'})
    f_bold  = workbook.add_format({'bold': True})
    f_head  = workbook.add_format({'bold': True, 'bg_color': '#E8F5E9', 'border': 1})
    f_head2 = workbook.add_format({'bold': True, 'bg_color': '#F3F4F6', 'border': 1})
    f_cell  = workbook.add_format({'border': 1})
    f_num   = workbook.add_format({'border': 1, 'num_format': '#,##0.00'})
    f_dt    = workbook.add_format({'border': 1, 'num_format': 'dd/mm/yyyy hh:mm'})

    # Larguras padrão
    ws.set_column(0, 0, 34)  # Produto
    ws.set_column(1, 1, 18)  # Garrafas
    ws.set_column(2, 2, 18)  # Doses
    ws.set_column(3, 3, 20)  # Doses ML
    ws.set_column(4, 4, 22)  # Data
    ws.set_column(5, 5, 18)  # Usuário

    row = 0

    # Título e período
    ws.write(row, 0, f"Relatório de Contagem — {restaurante.nome}", f_title); row += 1
    if use_range:
        label_modo = "Dia Operacional" if modo != 'calendario' else "Calendário"
        periodo_txt = f"{label_modo}: {timezone.localtime(inicio).strftime('%d/%m/%Y %H:%M')} → {timezone.localtime(fim).strftime('%d/%m/%Y %H:%M')}"
        ws.write(row, 0, periodo_txt, f_sub); row += 1
    row += 1

    # ===== Totais do restaurante =====
    ws.write(row, 0, "Total por Produto no Restaurante", f_bold); row += 1
    headers_total = ["Produto", "Total de Garrafas", "Total de Doses", "Total de Doses (ML)"]
    for col, h in enumerate(headers_total):
        ws.write(row, col, h, f_head)
    row += 1

    for item in somatorio_total.values():
        ws.write(row, 0, item['produto'].nome, f_cell)
        ws.write(row, 1, item['garrafas'], f_cell)
        ws.write(row, 2, item['doses'], f_num)
        ws.write(row, 3, item['doses'] * 50, f_num)
        row += 1

    row += 2  # espaço

    # ===== Por bar =====
    for bar_nome, contagens in dados_por_bar.items():
        ws.write(row, 0, f"Bar: {bar_nome}", f_bold); row += 1

        headers = ["Produto", "Garrafas", "Doses", "Doses (ML)", "Data da Contagem", "Usuário"]
        for col, h in enumerate(headers):
            ws.write(row, col, h, f_head2)
        row += 1

        for c in contagens:
            doses = float(c.quantidade_doses_restantes or 0)
            ws.write(row, 0, c.produto.nome, f_cell)
            ws.write(row, 1, c.quantidade_garrafas_cheias, f_cell)
            ws.write(row, 2, doses, f_num)
            ws.write(row, 3, doses * 50, f_num)

            # Data local (sem tz) para o Excel
            dt_local = timezone.localtime(c.data_contagem)
            ws.write_datetime(row, 4, dt_local.replace(tzinfo=None), f_dt)

            username = getattr(c.usuario, 'username', '-')
            ws.write(row, 5, username, f_cell)
            row += 1

        row += 2  # espaço entre bares

    workbook.close()
    output.seek(0)

    sufixo = "atual"
    if use_range:
        sufixo = f"{('oper' if modo != 'calendario' else 'cal')}_{inicio.strftime('%Y%m%d')}"
    filename = f"relatorio_contagem_{sufixo}_{restaurante.nome.replace(' ', '_')}.xlsx"

    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename={filename}'
    return resp



DOSE_ML = Decimal('50')

def _auto_fit_columns(ws, min_width=10, max_width=45):
    """Autoajusta a largura das colunas com limites razoáveis."""
    for col_cells in ws.columns:
        length = 0
        for cell in col_cells:
            v = cell.value
            v = '' if v is None else str(v)
            length = max(length, len(v))
        col = get_column_letter(col_cells[0].column)
        # um pouquinho de folga
        ws.column_dimensions[col].width = max(min_width, min(max_width, length + 2))

def _apply_header_style(row):
    """Aplica estilo de cabeçalho (cor, bold, centralizado)."""
    header_fill = PatternFill("solid", fgColor="E0E7FF")  # indigo-100
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))
    for c in row:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = border

def _apply_body_borders(ws, start_row, end_row, start_col, end_col):
    """Borda fininha nas células do corpo da tabela."""
    border = Border(left=Side(style="thin", color="EEEEEE"),
                    right=Side(style="thin", color="EEEEEE"),
                    top=Side(style="thin", color="EEEEEE"),
                    bottom=Side(style="thin", color="EEEEEE"))
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = border

@login_required
def exportar_diferenca_contagens_excel(request):
    bar_id = request.session.get('bar_id')
    if not bar_id:
        return HttpResponse("Nenhum bar selecionado.", status=400)

    bar_atual = get_object_or_404(Bar, id=bar_id)
    restaurante = bar_atual.restaurante
    bares = Bar.objects.filter(restaurante=restaurante).order_by('nome')

    # ===== Reaproveita a lógica do relatório para montar os dados =====
    dados_por_bar = {}
    somatorio_total = defaultdict(lambda: {
        'produto': None,
        'diff_garrafas': Decimal('0'),
        'diff_doses': Decimal('0'),
    })

    for bar in bares:
        contagens = (
            ContagemBar.objects
            .filter(bar=bar)
            .select_related('produto', 'usuario')
            .order_by('-data_contagem')
        )

        duas_ultimas_por_produto = {}
        for c in contagens:
            pid = c.produto_id
            if pid not in duas_ultimas_por_produto:
                duas_ultimas_por_produto[pid] = [c]
            elif len(duas_ultimas_por_produto[pid]) == 1:
                duas_ultimas_por_produto[pid].append(c)

        linhas_bar = []
        for pid, lista in duas_ultimas_por_produto.items():
            ultimo = lista[0]
            penultimo = lista[1] if len(lista) > 1 else None

            u_g = Decimal(ultimo.quantidade_garrafas_cheias or 0)
            u_d = Decimal(ultimo.quantidade_doses_restantes or 0)

            if penultimo:
                p_g = Decimal(penultimo.quantidade_garrafas_cheias or 0)
                p_d = Decimal(penultimo.quantidade_doses_restantes or 0)
                diff_g = u_g - p_g
                diff_d = u_d - p_d
            else:
                p_g = None
                p_d = None
                diff_g = None
                diff_d = None

            if diff_g is not None and diff_d is not None:
                somatorio_total[pid]['produto'] = ultimo.produto
                somatorio_total[pid]['diff_garrafas'] += diff_g
                somatorio_total[pid]['diff_doses'] += diff_d

            linhas_bar.append({
                'bar': bar.nome,
                'produto': ultimo.produto,
                'u_g': u_g, 'u_d': u_d,
                'p_g': p_g, 'p_d': p_d,
                'diff_g': diff_g, 'diff_d': diff_d,
                'data_p': (penultimo.data_contagem if penultimo else None),
                'user_p': (penultimo.usuario.username if penultimo else None),
                'data_u': ultimo.data_contagem,
                'user_u': ultimo.usuario.username,
            })

        linhas_bar.sort(key=lambda x: x['produto'].nome.lower())
        dados_por_bar[bar.nome] = linhas_bar

    somatorio_total = dict(sorted(
        somatorio_total.items(),
        key=lambda kv: kv[1]['produto'].nome.lower() if kv[1]['produto'] else ''
    ))

    # ===== Monta o Excel =====
    wb = Workbook()

    # Metadados / primeira sheet: Consolidado
    ws1 = wb.active
    ws1.title = "Consolidado"

    # Título e info
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    titulo = ws1.cell(row=1, column=1, value=f"Diferenças (Última − Penúltima) — {restaurante.nome}")
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal="left", vertical="center")

    now = timezone.localtime(timezone.now())
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    info = ws1.cell(row=2, column=1, value=f"Gerado em {now.strftime('%d/%m/%Y %H:%M')}")
    info.font = Font(italic=True, size=11)

    # Cabeçalho
    headers1 = ["Produto", "Garrafas", "Doses", "Doses (ML)"]
    ws1.append([])
    ws1.append(headers1)
    _apply_header_style(ws1[4])

    # Linhas do consolidado
    first_data_row = 5
    r = first_data_row
    for item in somatorio_total.values():
        prod = item['produto'].nome if item['produto'] else ""
        diff_g = item['diff_garrafas']
        diff_d = item['diff_doses']
        diff_ml = (diff_d * DOSE_ML) if diff_d is not None else None

        ws1.cell(row=r, column=1, value=prod)
        c2 = ws1.cell(row=r, column=2, value=float(diff_g) if diff_g is not None else None)
        c3 = ws1.cell(row=r, column=3, value=float(diff_d) if diff_d is not None else None)
        c4 = ws1.cell(row=r, column=4, value=float(diff_ml) if diff_ml is not None else None)

        c2.number_format = "0"
        c3.number_format = "0.00"
        c4.number_format = "0.00"
        r += 1

    if r > first_data_row:
        _apply_body_borders(ws1, first_data_row, r - 1, 1, 4)
        ws1.freeze_panes = "A5"

        # Escala de cores (vermelho → branco → verde) nas diferenças
        for col in [2, 3, 4]:
            col_letter = get_column_letter(col)
            ws1.conditional_formatting.add(
                f"{col_letter}{first_data_row}:{col_letter}{r-1}",
                ColorScaleRule(start_type='num', start_value=-1, start_color='FCA5A5',  # red-300
                               mid_type='num', mid_value=0, mid_color='FFFFFF',
                               end_type='num', end_value=1, end_color='86EFAC')   # green-300
            )

    _auto_fit_columns(ws1)

    # Segunda sheet: Detalhado por bar
    ws2 = wb.create_sheet(title="Detalhado")
    ws2.append([f"Diferenças por Produto (Penúltima → Última) — {restaurante.nome}"])
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws2.append([f"Gerado em {now.strftime('%d/%m/%Y %H:%M')}"])
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws2.cell(row=2, column=1).font = Font(italic=True, size=11)

    headers2 = [
        "Bar", "Produto",
        "Penúltima (Garrafas)", "Penúltima (Doses)",
        "Última (Garrafas)", "Última (Doses)",
        "Diferença (Garrafas)", "Diferença (Doses)", "Diferença (Doses ML)",
        "Data Penúltima", "Usuário Penúltima",
        "Data Última", "Usuário Última",
    ]
    ws2.append([])
    ws2.append(headers2)
    _apply_header_style(ws2[4])

    first_data_row2 = 5
    r2 = first_data_row2
    for bar_nome, linhas in dados_por_bar.items():
        for linha in linhas:
            diff_ml = (linha['diff_d'] * DOSE_ML) if linha['diff_d'] is not None else None
            ws2.cell(row=r2, column=1, value=bar_nome)
            ws2.cell(row=r2, column=2, value=linha['produto'].nome)

            c_pg = ws2.cell(row=r2, column=3, value=float(linha['p_g']) if linha['p_g'] is not None else None)
            c_pd = ws2.cell(row=r2, column=4, value=float(linha['p_d']) if linha['p_d'] is not None else None)
            c_ug = ws2.cell(row=r2, column=5, value=float(linha['u_g']))
            c_ud = ws2.cell(row=r2, column=6, value=float(linha['u_d']))
            c_dg = ws2.cell(row=r2, column=7, value=float(linha['diff_g']) if linha['diff_g'] is not None else None)
            c_dd = ws2.cell(row=r2, column=8, value=float(linha['diff_d']) if linha['diff_d'] is not None else None)
            c_ml = ws2.cell(row=r2, column=9, value=float(diff_ml) if diff_ml is not None else None)

            for c in (c_pg, c_ug, c_dg):
                if c is not None:
                    c.number_format = "0"
            for c in (c_pd, c_ud, c_dd, c_ml):
                if c is not None:
                    c.number_format = "0.00"

            ws2.cell(row=r2, column=10, value=linha['data_p'].strftime('%d/%m/%Y %H:%M') if linha['data_p'] else None)
            ws2.cell(row=r2, column=11, value=linha['user_p'] if linha['user_p'] else None)
            ws2.cell(row=r2, column=12, value=linha['data_u'].strftime('%d/%m/%Y %H:%M'))
            ws2.cell(row=r2, column=13, value=linha['user_u'])

            r2 += 1

    if r2 > first_data_row2:
        _apply_body_borders(ws2, first_data_row2, r2 - 1, 1, 13)
        ws2.freeze_panes = "A5"

        # Escala de cores para as 3 colunas de diferença (garrafas, doses, ml)
        for col in [7, 8, 9]:
            col_letter = get_column_letter(col)
            ws2.conditional_formatting.add(
                f"{col_letter}{first_data_row2}:{col_letter}{r2-1}",
                ColorScaleRule(start_type='num', start_value=-1, start_color='FCA5A5',
                               mid_type='num', mid_value=0, mid_color='FFFFFF',
                               end_type='num', end_value=1, end_color='86EFAC')
            )

    _auto_fit_columns(ws2)

    # ===== Resposta HTTP =====
    from django.utils.text import slugify
    filename = f"dif-contagens-{slugify(restaurante.nome)}-{now.strftime('%Y%m%d-%H%M')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




# ===== helpers simples p/ Excel =====
_THIN = Side(style="thin", color="DDDDDD")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_HEADER = PatternFill("solid", fgColor="F2F2F2")


def _x_style_header(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _FILL_HEADER
        cell.border = _BORDER_ALL


def _x_style_table(ws, start_row=2):
    if ws.max_row < start_row:
        return
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = _BORDER_ALL


def _x_auto_fit(ws, extra=2, max_width=60):
    widths = defaultdict(int)
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            val_str = "" if val is None else str(val)
            widths[idx] = max(widths[idx], len(val_str))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(w + extra, max_width)


def _parse_date_or_none(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


@login_required
def exportar_relatorio_perdas_excel(request):
    # 🔒 mesma permissão dos relatórios
    if not PermissaoPagina.objects.filter(user=request.user, nome_pagina='relatorios').exists():
        return HttpResponse("Sem permissão para exportar.", status=403)

    restaurante_id = request.session.get('restaurante_id')

    # ===== filtros (iguais ao relatório) =====
    bar_id = (request.GET.get('bar') or "").strip()
    q = (request.GET.get('q') or "").strip()                  # nome/código do produto
    motivo = (request.GET.get('motivo') or "").strip()        # código do motivo
    di_param = request.GET.get('data_inicio')
    df_param = request.GET.get('data_fim')

    hoje = timezone.localdate()
    data_inicio = _parse_date_or_none(di_param) or hoje.replace(day=1)
    data_fim = _parse_date_or_none(df_param) or hoje

    qs = (PerdaProduto.objects
          .select_related("bar", "produto", "usuario", "restaurante")
          .filter(data_registro__date__range=(data_inicio, data_fim))
          .order_by("data_registro"))

    if restaurante_id:
        qs = qs.filter(restaurante_id=restaurante_id)
    if bar_id:
        qs = qs.filter(bar_id=bar_id)
    if q:
        qs = qs.filter(Q(produto__nome__icontains=q) | Q(produto__codigo__icontains=q))
    if motivo:
        qs = qs.filter(motivo=motivo)

    # ===== agregações =====
    por_produto = defaultdict(lambda: {"garrafas": 0, "doses": 0})
    por_bar = defaultdict(lambda: {"garrafas": 0, "doses": 0})
    por_motivo = defaultdict(lambda: {"registros": 0, "garrafas": 0, "doses": 0})

    total_registros = 0
    total_garrafas = 0
    total_doses = 0

    for p in qs:
        total_registros += 1
        g = int(p.garrafas or 0)
        d = int(p.doses or 0)
        total_garrafas += g
        total_doses += d

        # produto
        nome_prod = f"[{getattr(p.produto, 'codigo', '')}] {p.produto.nome}" if getattr(p.produto, 'codigo', None) else p.produto.nome
        por_produto[nome_prod]["garrafas"] += g
        por_produto[nome_prod]["doses"] += d

        # bar
        por_bar[p.bar.nome]["garrafas"] += g
        por_bar[p.bar.nome]["doses"] += d

        # motivo
        por_motivo[p.motivo]["registros"] += 1
        por_motivo[p.motivo]["garrafas"] += g
        por_motivo[p.motivo]["doses"] += d

    # map de rótulo do motivo
    motivos_map = dict(PerdaProduto.MOTIVOS)

    # ===== monta workbook =====
    wb = Workbook()

    # --- Resumo ---
    ws = wb.active
    ws.title = "Resumo"

    titulo = f"Relatório de Perdas — {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    ws.merge_cells("A1:D1")
    ws["A1"] = titulo
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # filtros ativos
    filtros = []
    if restaurante_id:
        filtros.append(f"Restaurante ID: {restaurante_id}")
    if bar_id:
        filtros.append(f"Bar ID: {bar_id}")
    if q:
        filtros.append(f"Produto: {q}")
    if motivo:
        filtros.append(f"Motivo: {motivos_map.get(motivo, motivo)}")

    ws.append([])
    ws.append(["Filtros aplicados", (", ".join(filtros) or "Nenhum")])
    ws.append(["Total de registros", total_registros])
    ws.append(["Total de garrafas", total_garrafas])
    ws.append(["Total de doses", total_doses])

    # negrito nas métricas
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")

    _x_auto_fit(ws)

    # por motivo
    ws_m = wb.create_sheet("Por Motivo")
    ws_m.append(["Motivo", "Registros", "Garrafas", "Doses"])
    _x_style_header(ws_m, 1)

    for cod, d in sorted(por_motivo.items(), key=lambda kv: motivos_map.get(kv[0], kv[0])):
        ws_m.append([motivos_map.get(cod, cod), d["registros"], d["garrafas"], d["doses"]])

    # total linha final
    if ws_m.max_row > 1:
        ws_m.append(["TOTAL",
                     sum(d["registros"] for d in por_motivo.values()),
                     sum(d["garrafas"] for d in por_motivo.values()),
                     sum(d["doses"] for d in por_motivo.values())])
        for c in range(1, ws_m.max_column + 1):
            cell = ws_m.cell(row=ws_m.max_row, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EFEFEF")

    _x_style_table(ws_m, start_row=2)
    for col in ("B", "C", "D"):
        for cell in ws_m[col][1:]:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0"
    _x_auto_fit(ws_m)

    # --- Por Produto ---
    ws_p = wb.create_sheet("Por Produto")
    ws_p.append(["Produto", "Garrafas", "Doses"])
    _x_style_header(ws_p, 1)

    for nome, d in sorted(por_produto.items(), key=lambda kv: kv[0].lower()):
        ws_p.append([nome, d["garrafas"], d["doses"]])

    if ws_p.max_row > 1:
        ws_p.append(["TOTAL",
                     sum(v["garrafas"] for v in por_produto.values()),
                     sum(v["doses"] for v in por_produto.values())])
        last = ws_p.max_row
        for c in range(1, ws_p.max_column + 1):
            cell = ws_p.cell(row=last, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EFEFEF")

    _x_style_table(ws_p, start_row=2)
    for col in ("B", "C"):
        for cell in ws_p[col][1:]:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0"
    _x_auto_fit(ws_p)

    # --- Por Bar ---
    ws_b = wb.create_sheet("Por Bar")
    ws_b.append(["Bar", "Garrafas", "Doses"])
    _x_style_header(ws_b, 1)

    for nome_bar, d in sorted(por_bar.items(), key=lambda kv: kv[0].lower()):
        ws_b.append([nome_bar, d["garrafas"], d["doses"]])

    if ws_b.max_row > 1:
        ws_b.append(["TOTAL",
                     sum(v["garrafas"] for v in por_bar.values()),
                     sum(v["doses"] for v in por_bar.values())])
        last = ws_b.max_row
        for c in range(1, ws_b.max_column + 1):
            cell = ws_b.cell(row=last, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EFEFEF")

    _x_style_table(ws_b, start_row=2)
    for col in ("B", "C"):
        for cell in ws_b[col][1:]:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0"
    _x_auto_fit(ws_b)

    # --- Detalhe ---
    ws_d = wb.create_sheet("Detalhe")
    ws_d.append(["Data", "Hora", "Restaurante", "Bar", "Produto", "Garrafas", "Doses", "Motivo", "Usuário", "Observação"])
    _x_style_header(ws_d, 1)

    for p in qs:
        data_local = timezone.localtime(p.data_registro)
        prod_label = f"[{getattr(p.produto, 'codigo', '')}] {p.produto.nome}" if getattr(p.produto, 'codigo', None) else p.produto.nome
        ws_d.append([
            data_local.date().strftime("%d/%m/%Y"),
            data_local.time().strftime("%H:%M"),
            getattr(p.restaurante, "nome", ""),
            p.bar.nome,
            prod_label,
            int(p.garrafas or 0),
            int(p.doses or 0),
            motivos_map.get(p.motivo, p.motivo),
            getattr(p.usuario, "username", ""),
            p.observacao or "",
        ])

    _x_style_table(ws_d, start_row=2)
    for col in ("F", "G"):
        for cell in ws_d[col][1:]:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0"
    _x_auto_fit(ws_d)

    # ===== resposta =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nome_arquivo = f"relatorio_perdas_{data_inicio:%Y-%m-%d}_a_{data_fim:%Y-%m-%d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response






#                                                                                  GRAFICOS


